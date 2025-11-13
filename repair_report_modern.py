#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
维修单工具 - Modern UI Edition v2.0.0
RepoPrompt-inspired glassmorphism design with semi-transparent effects
"""

import customtkinter as ctk
from tkinter import filedialog, messagebox, scrolledtext
import tkinter as tk
from PIL import Image, ImageTk, ImageFilter, ImageDraw
import os
import json
from datetime import datetime
from pathlib import Path
import tempfile
import uuid
import platform
import sys
import glob
import re

# Set appearance mode and color theme
ctk.set_appearance_mode("light")  # Light mode only
ctk.set_default_color_theme("blue")  # Themes: "blue", "green", "dark-blue"

# Excel export
try:
    import openpyxl
    from openpyxl.drawing import image as xl_image
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# PDF export
try:
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RL_Image, PageBreak, KeepTogether
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch, mm
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

# Drag and drop
DRAG_DROP_AVAILABLE = False
DND_FILES = None
try:
    from tkinterdnd2 import DND_FILES, TkinterDnD
    DRAG_DROP_AVAILABLE = True
except ImportError:
    pass


class ModernRepairTool(ctk.CTk):
    """Modern repair report tool with glassmorphism UI"""

    def __init__(self):
        super().__init__()

        # Window configuration
        self.title("维修单工具 Modern Edition v2.0")
        self.geometry("1600x1000")

        # Set window transparency for glassmorphism effect
        try:
            self.attributes('-alpha', 0.96)  # Slightly more transparent for light theme
        except:
            pass

        # Data model
        self.project_title_var = ctk.StringVar(value="")
        self.items = []
        self.current_item_id = 0
        self.selected_item_index = None
        self.max_images_per_row = 1

        # Caches
        self.image_cache = {}
        self.thumbnail_cache = {}

        # Debug logs
        self.debug_logs = []
        self.max_debug_logs = 500

        # Color scheme - Light and transparent
        self.colors = {
            'bg_primary': '#F8F9FA',      # Light gray background
            'bg_secondary': '#FFFFFF',    # White
            'bg_tertiary': '#E9ECEF',     # Slightly gray
            'accent': '#0D6EFD',          # Blue accent
            'accent_hover': '#0B5ED7',    # Darker blue
            'text_primary': '#212529',    # Dark text
            'text_secondary': '#6C757D',  # Gray text
            'border': '#DEE2E6',          # Light border
            'success': '#198754',         # Green
            'warning': '#FFC107',         # Orange/Yellow
            'error': '#DC3545',           # Red
            'glass': 'rgba(255, 255, 255, 0.8)'  # Light glassmorphism
        }

        # Initialize UI
        self.create_menu()  # Add menu first
        self.setup_ui()
        self.setup_drag_drop()
        self.bind_shortcuts()  # Add keyboard shortcuts

    def setup_ui(self):
        """Setup the modern UI layout"""
        # Configure grid
        self.grid_columnconfigure(0, weight=0)  # Sidebar
        self.grid_columnconfigure(1, weight=1)  # Main content
        self.grid_rowconfigure(0, weight=0)     # Top bar
        self.grid_rowconfigure(1, weight=1)     # Content area
        self.grid_rowconfigure(2, weight=0)     # Status bar

        # Create sections
        self.create_top_bar()
        self.create_sidebar()
        self.create_main_area()
        self.create_status_bar()

    def create_top_bar(self):
        """Create modern top bar with title and controls"""
        top_frame = ctk.CTkFrame(
            self,
            height=80,
            corner_radius=0,
            fg_color=self.colors['bg_secondary']
        )
        top_frame.grid(row=0, column=0, columnspan=2, sticky="ew", padx=0, pady=0)
        top_frame.grid_columnconfigure(1, weight=1)

        # App icon and title
        title_frame = ctk.CTkFrame(top_frame, fg_color="transparent")
        title_frame.grid(row=0, column=0, sticky="w", padx=20, pady=15)

        ctk.CTkLabel(
            title_frame,
            text="🔧",
            font=ctk.CTkFont(size=32)
        ).pack(side="left", padx=(0, 10))

        ctk.CTkLabel(
            title_frame,
            text="维修单工具",
            font=ctk.CTkFont(size=24, weight="bold"),
            text_color=self.colors['text_primary']
        ).pack(side="left")

        ctk.CTkLabel(
            title_frame,
            text="Modern Edition",
            font=ctk.CTkFont(size=12),
            text_color=self.colors['text_secondary']
        ).pack(side="left", padx=(10, 0))

        # Project title input (centered)
        title_input_frame = ctk.CTkFrame(top_frame, fg_color="transparent")
        title_input_frame.grid(row=0, column=1, sticky="ew", padx=40, pady=15)

        title_label_frame = ctk.CTkFrame(title_input_frame, fg_color="transparent")
        title_label_frame.pack(fill="x", pady=(0, 5))

        ctk.CTkLabel(
            title_label_frame,
            text="项目标题",
            font=ctk.CTkFont(size=12),
            text_color=self.colors['text_secondary']
        ).pack(side="left")

        ctk.CTkButton(
            title_label_frame,
            text="快速填充",
            width=80,
            height=24,
            corner_radius=6,
            font=ctk.CTkFont(size=11),
            fg_color=self.colors['accent'],
            hover_color=self.colors['accent_hover'],
            command=self.quick_fill_title
        ).pack(side="right")

        self.title_entry = ctk.CTkEntry(
            title_input_frame,
            textvariable=self.project_title_var,
            placeholder_text="输入项目标题...",
            height=40,
            font=ctk.CTkFont(size=14),
            corner_radius=10
        )
        self.title_entry.pack(fill="x")

        # Action buttons (right side)
        action_frame = ctk.CTkFrame(top_frame, fg_color="transparent")
        action_frame.grid(row=0, column=2, sticky="e", padx=20, pady=15)

        # Export buttons with icons
        if EXCEL_AVAILABLE:
            self.create_modern_button(
                action_frame,
                "📊 Excel",
                self.export_excel,
                "left",
                success=False
            )

        if PDF_AVAILABLE:
            self.create_modern_button(
                action_frame,
                "📄 PDF",
                self.export_pdf,
                "left",
                success=False
            )

        # Settings button
        self.create_modern_button(
            action_frame,
            "⚙️",
            self.show_settings,
            "left",
            width=50
        )

    def create_modern_button(self, parent, text, command, side="left", width=120, success=False):
        """Create a modern glassmorphic button"""
        btn = ctk.CTkButton(
            parent,
            text=text,
            command=command,
            width=width,
            height=40,
            corner_radius=10,
            font=ctk.CTkFont(size=13, weight="bold"),
            fg_color=self.colors['success'] if success else self.colors['accent'],
            hover_color=self.colors['accent_hover']
        )
        btn.pack(side=side, padx=5)
        return btn

    def create_sidebar(self):
        """Create modern sidebar with project list"""
        sidebar = ctk.CTkFrame(
            self,
            width=380,
            corner_radius=0,
            fg_color=self.colors['bg_secondary']
        )
        sidebar.grid(row=1, column=0, sticky="nsew", padx=0, pady=0)
        sidebar.grid_propagate(False)

        # Sidebar header
        header = ctk.CTkFrame(sidebar, fg_color="transparent", height=60)
        header.pack(fill="x", padx=20, pady=(20, 10))

        ctk.CTkLabel(
            header,
            text="维修项目",
            font=ctk.CTkFont(size=18, weight="bold"),
            text_color=self.colors['text_primary']
        ).pack(side="left")

        # Add button
        # Control buttons
        btn_group = ctk.CTkFrame(header, fg_color="transparent")
        btn_group.pack(side="right")

        ctk.CTkButton(
            btn_group,
            text="↑",
            width=32,
            height=32,
            corner_radius=8,
            font=ctk.CTkFont(size=14, weight="bold"),
            fg_color=self.colors['bg_tertiary'],
            hover_color=self.colors['border'],
            command=self.move_item_up
        ).pack(side="left", padx=2)

        ctk.CTkButton(
            btn_group,
            text="↓",
            width=32,
            height=32,
            corner_radius=8,
            font=ctk.CTkFont(size=14, weight="bold"),
            fg_color=self.colors['bg_tertiary'],
            hover_color=self.colors['border'],
            command=self.move_item_down
        ).pack(side="left", padx=2)

        ctk.CTkButton(
            btn_group,
            text="+ 添加",
            width=80,
            height=32,
            corner_radius=8,
            font=ctk.CTkFont(size=12, weight="bold"),
            fg_color=self.colors['accent'],
            hover_color=self.colors['accent_hover'],
            command=self.add_item
        ).pack(side="left", padx=2)

        # Search bar
        self.search_entry = ctk.CTkEntry(
            sidebar,
            placeholder_text="🔍 搜索项目...",
            height=36,
            corner_radius=8
        )
        self.search_entry.pack(fill="x", padx=20, pady=(0, 15))

        # Scrollable frame for items
        self.items_scroll = ctk.CTkScrollableFrame(
            sidebar,
            fg_color="transparent",
            corner_radius=0
        )
        self.items_scroll.pack(fill="both", expand=True, padx=15, pady=0)

        # Stats at bottom
        stats_frame = ctk.CTkFrame(sidebar, fg_color=self.colors['bg_tertiary'], height=80)
        stats_frame.pack(fill="x", padx=15, pady=15, side="bottom")

        self.stats_label = ctk.CTkLabel(
            stats_frame,
            text="0 项目 • 0 图片",
            font=ctk.CTkFont(size=13),
            text_color=self.colors['text_secondary']
        )
        self.stats_label.pack(pady=15)

        # Action buttons
        btn_frame = ctk.CTkFrame(sidebar, fg_color="transparent")
        btn_frame.pack(fill="x", padx=20, pady=(0, 15), side="bottom")

        ctk.CTkButton(
            btn_frame,
            text="📁 打开",
            command=self.open_project,
            height=36,
            corner_radius=8,
            fg_color=self.colors['bg_tertiary'],
            hover_color=self.colors['border']
        ).pack(side="left", expand=True, padx=(0, 5))

        ctk.CTkButton(
            btn_frame,
            text="💾 保存",
            command=self.save_project,
            height=36,
            corner_radius=8,
            fg_color=self.colors['bg_tertiary'],
            hover_color=self.colors['border']
        ).pack(side="right", expand=True, padx=(5, 0))

    def create_main_area(self):
        """Create main content area for images"""
        main_frame = ctk.CTkFrame(
            self,
            corner_radius=0,
            fg_color=self.colors['bg_primary']
        )
        main_frame.grid(row=1, column=1, sticky="nsew", padx=0, pady=0)

        # Content header
        content_header = ctk.CTkFrame(main_frame, fg_color="transparent", height=70)
        content_header.pack(fill="x", padx=30, pady=(20, 10))

        # Item description area
        desc_frame = ctk.CTkFrame(content_header, fg_color="transparent")
        desc_frame.pack(fill="x")

        ctk.CTkLabel(
            desc_frame,
            text="项目描述",
            font=ctk.CTkFont(size=16, weight="bold"),
            text_color=self.colors['text_primary']
        ).pack(anchor="w", pady=(0, 5))

        self.description_entry = ctk.CTkEntry(
            desc_frame,
            placeholder_text="输入维修项目描述...",
            height=40,
            font=ctk.CTkFont(size=13),
            corner_radius=8
        )
        self.description_entry.pack(fill="x", pady=(0, 10))
        self.description_entry.bind('<KeyRelease>', self.on_description_change)

        # Image controls
        img_control_frame = ctk.CTkFrame(content_header, fg_color="transparent")
        img_control_frame.pack(fill="x")

        ctk.CTkLabel(
            img_control_frame,
            text="图片管理",
            font=ctk.CTkFont(size=14),
            text_color=self.colors['text_secondary']
        ).pack(side="left")

        # Image action buttons
        ctk.CTkButton(
            img_control_frame,
            text="📸 添加图片",
            command=self.add_images,
            height=32,
            corner_radius=8,
            font=ctk.CTkFont(size=12),
            fg_color=self.colors['accent'],
            hover_color=self.colors['accent_hover']
        ).pack(side="right", padx=5)

        ctk.CTkButton(
            img_control_frame,
            text="📁 批量添加",
            command=self.batch_add_images,
            height=32,
            corner_radius=8,
            font=ctk.CTkFont(size=12),
            fg_color=self.colors['bg_tertiary'],
            hover_color=self.colors['border']
        ).pack(side="right", padx=5)

        # Scrollable image gallery
        self.image_gallery = ctk.CTkScrollableFrame(
            main_frame,
            fg_color="transparent",
            corner_radius=0
        )
        self.image_gallery.pack(fill="both", expand=True, padx=30, pady=(0, 20))

        # Configure grid for image gallery
        for i in range(4):
            self.image_gallery.grid_columnconfigure(i, weight=1)

        # Drop zone indicator
        self.drop_zone = ctk.CTkFrame(
            self.image_gallery,
            fg_color=self.colors['bg_secondary'],
            corner_radius=15,
            border_width=2,
            border_color=self.colors['border']
        )
        self.drop_zone.pack(fill="both", expand=True, padx=20, pady=50)

        drop_content = ctk.CTkFrame(self.drop_zone, fg_color="transparent")
        drop_content.place(relx=0.5, rely=0.5, anchor="center")

        ctk.CTkLabel(
            drop_content,
            text="📁",
            font=ctk.CTkFont(size=64)
        ).pack(pady=(0, 15))

        ctk.CTkLabel(
            drop_content,
            text="拖拽图片到这里" if DRAG_DROP_AVAILABLE else "点击'添加图片'按钮",
            font=ctk.CTkFont(size=18, weight="bold"),
            text_color=self.colors['text_primary']
        ).pack()

        ctk.CTkLabel(
            drop_content,
            text="或点击'添加图片'按钮选择文件",
            font=ctk.CTkFont(size=13),
            text_color=self.colors['text_secondary']
        ).pack(pady=(5, 0))

    def create_status_bar(self):
        """Create modern status bar"""
        status_frame = ctk.CTkFrame(
            self,
            height=40,
            corner_radius=0,
            fg_color=self.colors['bg_secondary']
        )
        status_frame.grid(row=2, column=0, columnspan=2, sticky="ew", padx=0, pady=0)

        self.status_label = ctk.CTkLabel(
            status_frame,
            text="就绪",
            font=ctk.CTkFont(size=11),
            text_color=self.colors['text_secondary']
        )
        self.status_label.pack(side="left", padx=20)

        # Feature indicators
        features = []
        if DRAG_DROP_AVAILABLE:
            features.append("✓ 拖拽")
        if EXCEL_AVAILABLE:
            features.append("✓ Excel")
        if PDF_AVAILABLE:
            features.append("✓ PDF")

        feature_text = " • ".join(features)
        ctk.CTkLabel(
            status_frame,
            text=feature_text,
            font=ctk.CTkFont(size=11),
            text_color=self.colors['success']
        ).pack(side="right", padx=20)

    def setup_drag_drop(self):
        """Setup drag and drop if available"""
        if not DRAG_DROP_AVAILABLE:
            return

        try:
            # Load the tkdnd package
            self.tk.eval('package require tkdnd')
            self._log_debug("✓ tkdnd 包已加载")

            # Register the window as a drop target
            self.tk.call('::tkdnd::drop_target', 'register', self._w, DND_FILES)
            self._log_debug(f"✓ 注册拖拽目标: {self._w}")

            # Store drop data in a Tcl variable that Python can read
            self.tk.setvar('drop_data', '')

            # Create Python callback for drop event
            def on_drop_handler():
                """Handler that gets called after Tcl stores the drop data"""
                try:
                    data = self.tk.getvar('drop_data')
                    self._log_debug(f"🎯 Drop handler 被调用，数据: {data}")
                    if data:
                        # Create a mock event object
                        class DropEvent:
                            def __init__(self, data):
                                self.data = data
                        self.on_drop(DropEvent(data))
                except Exception as e:
                    self._log_debug(f"❌ Drop handler 异常: {e}")
                    import traceback
                    traceback.print_exc()

            # Register the handler and get the Tcl command name
            handler_cmd = self.register(on_drop_handler)
            self._log_debug(f"✓ 注册 Python 回调，Tcl 命令: {handler_cmd}")

            # Use a simpler Tcl binding that stores data and calls Python
            # Must register handler BEFORE using it in Tcl script
            self.tk.eval(f'''
                bind {self._w} <<Drop>> {{
                    set ::drop_data %D
                    after idle {handler_cmd}
                    return copy
                }}
                bind {self._w} <<DropEnter>> {{
                    return copy
                }}
                bind {self._w} <<DropPosition>> {{
                    return copy
                }}
                bind {self._w} <<DropLeave>> {{
                    # Do nothing
                }}
            ''')
            self._log_debug("✓ 使用简化的 Tcl 绑定（避免事件替换冲突）")

            # Also register on specific widgets after they're created
            self.after(200, self._register_widget_drops)

            print("✅ 拖拽功能设置成功")
            self._log_debug("✅ 拖拽功能完全初始化")
        except Exception as e:
            print(f"⚠️ 拖拽功能设置失败: {e}")
            self._log_debug(f"⚠️ 拖拽功能设置失败: {e}")
            import traceback
            traceback.print_exc()

    def _register_widget_drops(self):
        """Register drop on specific widgets after they're created"""
        try:
            # Create helper functions for widget drop registration
            def register_widget_drop(widget):
                try:
                    # Get the widget's window path name
                    widget_path = str(widget)
                    # Register using the low-level Tcl interface
                    self.tk.call('::tkdnd::drop_target', 'register', widget_path, DND_FILES)
                    # Bind the drop event using Python's bind
                    widget.bind('<<Drop>>', self.on_drop, '+')
                    self._log_debug(f"✓ 注册 widget 拖拽: {widget_path}")
                    return True
                except Exception as e:
                    self._log_debug(f"✗ 跳过 widget {widget}: {e}")
                    return False

            # Try to register on the main frame widgets
            widgets_to_register = []

            # Add drop zone
            if hasattr(self, 'drop_zone') and self.drop_zone:
                widgets_to_register.append(self.drop_zone)

            # Add image gallery
            if hasattr(self, 'image_gallery') and self.image_gallery:
                widgets_to_register.append(self.image_gallery)

            # Register each widget
            success_count = 0
            for widget in widgets_to_register:
                if register_widget_drop(widget):
                    success_count += 1

                # Also try to register internal frames
                if hasattr(widget, '_canvas'):
                    if register_widget_drop(widget._canvas):
                        success_count += 1
                if hasattr(widget, '_parent_canvas'):
                    if register_widget_drop(widget._parent_canvas):
                        success_count += 1

            if success_count > 0:
                print(f"✅ 成功注册 {success_count} 个拖拽区域")
                self._log_debug(f"✅ 成功注册 {success_count} 个拖拽区域")
        except Exception as e:
            print(f"⚠️ Widget 拖拽注册失败: {e}")
            self._log_debug(f"⚠️ Widget 拖拽注册失败: {e}")
            import traceback
            traceback.print_exc()

    def on_drop(self, event):
        """Handle drag and drop of image files"""
        if not DRAG_DROP_AVAILABLE:
            return

        try:
            raw_data = event.data or ""
            print(f"拖拽原始数据: {raw_data}")
            self._log_debug(f"📥 拖拽事件触发，原始数据: {raw_data}")

            # Parse file paths
            files = self._split_dnd_paths(raw_data)
            self._log_debug(f"📄 解析到 {len(files)} 个文件路径")

            # Filter for image files
            exts = ('.jpg', '.jpeg', '.png', '.gif', '.bmp', '.tiff', '.webp')
            image_files = [fp for fp in files if os.path.exists(fp) and fp.lower().endswith(exts)]
            self._log_debug(f"🖼️ 筛选到 {len(image_files)} 个有效图片文件")

            if not image_files:
                self._log_debug("⚠️ 未找到有效的图片文件")
                messagebox.showwarning("提示", "未找到有效的图片文件")
                return

            # Handle single or multiple images
            if len(image_files) == 1:
                # Add to current selected item
                if self.selected_item_index is not None and 0 <= self.selected_item_index < len(self.items):
                    img_path = image_files[0]
                    if img_path not in self.items[self.selected_item_index]['images']:
                        self.items[self.selected_item_index]['images'].append(img_path)
                        self.refresh_item_list()
                        self.display_item_images(self.selected_item_index)
                        self.update_stats()
                        self.set_status(f"✓ 已添加图片到项目 {self.selected_item_index + 1}")
                        self._log_debug(f"✅ 成功添加图片到项目 {self.selected_item_index + 1}")
                    else:
                        self._log_debug("⚠️ 图片已存在")
                        messagebox.showinfo("提示", "该图片已存在")
                else:
                    # No item selected, ask user
                    if messagebox.askyesno("提示", "当前没有选中项目\n是否创建新项目并添加图片？"):
                        self.add_item()
                        if self.selected_item_index is not None:
                            self.items[self.selected_item_index]['images'].append(image_files[0])
                            self.refresh_item_list()
                            self.display_item_images(self.selected_item_index)
                            self.update_stats()
            else:
                # Multiple images, show batch assign dialog
                self.show_batch_assign_dialog(image_files)

        except Exception as e:
            messagebox.showerror("拖拽错误", f"拖拽处理失败: {str(e)}")
            print(f"拖拽处理异常: {e}")
            import traceback
            traceback.print_exc()

    def _split_dnd_paths(self, raw):
        """Split drag-and-drop file paths (handles spaces and special chars)"""
        try:
            # Try tkinter's built-in splitlist first
            return [p.strip().strip('"').strip("'") for p in self.tk.splitlist(raw)]
        except Exception:
            # Fallback: manual parsing for paths with spaces
            # Handles formats like: {path1} {path2} or "path1" "path2"
            candidates = re.findall(r'\{([^}]*)\}|"([^"]*)"|\'([^\']*)\'|([^ \t\r\n]+)', raw or "")
            cleaned = []
            for groups in candidates:
                # groups is a tuple of matched groups, find the non-empty one
                path = next((g for g in groups if g), None)
                if path:
                    cleaned.append(path.strip())
            return cleaned

    def add_item(self):
        """Add new repair item"""
        self.current_item_id += 1
        new_item = {
            'id': self.current_item_id,
            'description': f"维修项目 {len(self.items) + 1}",
            'images': []
        }
        self.items.append(new_item)

        # Only create the new card instead of refreshing all
        idx = len(self.items) - 1
        self.create_item_card(idx, new_item)

        self.update_stats()
        self.set_status(f"✓ 已添加新项目")

        # Select the new item (optimized to not refresh all cards)
        self.select_item_optimized(idx)

    def refresh_item_list(self):
        """Refresh the sidebar item list"""
        # Clear existing items
        for widget in self.items_scroll.winfo_children():
            widget.destroy()

        # Add item cards
        for idx, item in enumerate(self.items):
            self.create_item_card(idx, item)

    def create_item_card(self, idx, item):
        """Create a modern card for an item"""
        # Determine if this card is selected
        is_selected = (idx == self.selected_item_index)
        border_width = 3 if is_selected else 2
        border_color = self.colors['accent'] if is_selected else self.colors['border']

        card = ctk.CTkFrame(
            self.items_scroll,
            fg_color=self.colors['bg_tertiary'],
            corner_radius=12,
            border_width=border_width,
            border_color=border_color
        )
        card.pack(fill="x", pady=8)

        # Card content
        content = ctk.CTkFrame(card, fg_color="transparent")
        content.pack(fill="x", padx=15, pady=12)

        # Item number badge
        badge = ctk.CTkLabel(
            content,
            text=str(idx + 1),
            font=ctk.CTkFont(size=16, weight="bold"),
            text_color=self.colors['accent'],
            width=35,
            height=35,
            fg_color=self.colors['bg_primary'],
            corner_radius=8
        )
        badge.pack(side="left", padx=(0, 12))

        # Item info
        info_frame = ctk.CTkFrame(content, fg_color="transparent")
        info_frame.pack(side="left", fill="x", expand=True)

        desc_text = item['description'][:30] + "..." if len(item['description']) > 30 else item['description']
        desc_label = ctk.CTkLabel(
            info_frame,
            text=desc_text,
            font=ctk.CTkFont(size=13),
            text_color=self.colors['text_primary'],
            anchor="w"
        )
        desc_label.pack(anchor="w")

        img_count = len(item.get('images', []))
        count_label = ctk.CTkLabel(
            info_frame,
            text=f"📷 {img_count} 张图片",
            font=ctk.CTkFont(size=11),
            text_color=self.colors['text_secondary'],
            anchor="w"
        )
        count_label.pack(anchor="w", pady=(2, 0))

        # Delete button
        del_btn = ctk.CTkButton(
            content,
            text="🗑",
            width=35,
            height=35,
            corner_radius=8,
            fg_color=self.colors['bg_primary'],
            hover_color=self.colors['error'],
            command=lambda: self.delete_item(idx)
        )
        del_btn.pack(side="right")

        # Bind click event to all widgets except delete button
        # This ensures clicking anywhere on the card selects the item
        click_handler = lambda e, i=idx: self.select_item_optimized(i)
        self._bind_click_recursive(card, click_handler, exclude=[del_btn])

    def _bind_click_recursive(self, widget, handler, exclude=None):
        """Recursively bind click event to widget and all children"""
        if exclude is None:
            exclude = []

        # Don't bind to excluded widgets (like delete button)
        if widget not in exclude:
            widget.bind("<Button-1>", handler, add="+")

            # Recursively bind to all children
            for child in widget.winfo_children():
                self._bind_click_recursive(child, handler, exclude)

    def update_item_card_text(self, idx):
        """Update only the text of a specific item card without rebuilding the entire list"""
        if idx < 0 or idx >= len(self.items):
            return

        # Find the card widget in the scrollable frame
        cards = self.items_scroll.winfo_children()
        if idx >= len(cards):
            return

        card = cards[idx]
        item = self.items[idx]

        # Find and update the description label
        # The structure is: card -> content -> info_frame -> desc_label
        try:
            content = card.winfo_children()[0]  # content frame
            info_frame = None

            # Find the info_frame (it's the one that's not a label and not a button)
            for child in content.winfo_children():
                if isinstance(child, ctk.CTkFrame) and child.cget('fg_color') == 'transparent':
                    info_frame = child
                    break

            if info_frame:
                # Update the description label (first label in info_frame)
                for child in info_frame.winfo_children():
                    if isinstance(child, ctk.CTkLabel):
                        desc_text = item['description'][:30] + "..." if len(item['description']) > 30 else item['description']
                        child.configure(text=desc_text)
                        break
        except Exception as e:
            # If updating fails, fallback to full refresh
            print(f"⚠️ 更新卡片文本失败，回退到完整刷新: {e}")
            self.refresh_item_list()

    def select_item(self, idx):
        """Select an item and display its images"""
        if 0 <= idx < len(self.items):
            self.selected_item_index = idx
            item = self.items[idx]

            # Update description
            self.description_entry.delete(0, "end")
            self.description_entry.insert(0, item['description'])

            # Display images
            self.display_item_images(idx)

            # Update card highlights
            self.refresh_item_list()

    def select_item_optimized(self, idx):
        """Select an item without refreshing all cards (optimized)"""
        if 0 <= idx < len(self.items):
            self.selected_item_index = idx
            item = self.items[idx]

            # Update description
            self.description_entry.delete(0, "end")
            self.description_entry.insert(0, item['description'])

            # Display images
            self.display_item_images(idx)

            # Update only the card highlights without rebuilding
            self.update_card_highlights()

    def update_card_highlights(self):
        """Update card highlights without rebuilding all cards"""
        for idx, widget in enumerate(self.items_scroll.winfo_children()):
            if idx < len(self.items):
                # Update border color based on selection
                if idx == self.selected_item_index:
                    widget.configure(border_color=self.colors['accent'], border_width=3)
                else:
                    widget.configure(border_color=self.colors['border'], border_width=2)

    def delete_item(self, idx):
        """Delete an item"""
        if 0 <= idx < len(self.items):
            if messagebox.askyesno("确认删除", f"确定要删除项目 {idx + 1} 吗？"):
                del self.items[idx]
                self.refresh_item_list()
                self.update_stats()
                self.clear_image_display()
                self.set_status(f"✓ 已删除项目")

    def on_description_change(self, event):
        """Handle description text change"""
        if self.selected_item_index is not None and 0 <= self.selected_item_index < len(self.items):
            new_desc = self.description_entry.get().strip()
            self.items[self.selected_item_index]['description'] = new_desc if new_desc else f"维修项目 {self.selected_item_index + 1}"
            # 只更新选中项的卡片文本，不刷新整个列表，避免跳闪
            self.update_item_card_text(self.selected_item_index)

    def add_images(self):
        """Add images to selected item"""
        if self.selected_item_index is None:
            messagebox.showwarning("提示", "请先选择一个项目")
            return

        file_paths = filedialog.askopenfilenames(
            title="选择图片文件",
            filetypes=[
                ("图片文件", "*.jpg *.jpeg *.png *.gif *.bmp *.tiff *.webp"),
                ("所有文件", "*.*")
            ]
        )

        if file_paths:
            item = self.items[self.selected_item_index]
            added = 0
            for path in file_paths:
                if path not in item['images']:
                    item['images'].append(path)
                    added += 1

            self.refresh_item_list()
            self.display_item_images(self.selected_item_index)
            self.update_stats()
            self.set_status(f"✓ 已添加 {added} 张图片")

    def batch_add_images(self):
        """Batch add images with assignment dialog"""
        if not self.items:
            if messagebox.askyesno("提示", "当前没有项目\n是否创建新项目？"):
                self.add_item()
            else:
                return

        file_paths = filedialog.askopenfilenames(
            title="选择图片文件",
            filetypes=[
                ("图片文件", "*.jpg *.jpeg *.png *.gif *.bmp *.tiff *.webp"),
                ("所有文件", "*.*")
            ]
        )

        if file_paths:
            self.show_batch_assign_dialog(list(file_paths))

    def show_batch_assign_dialog(self, file_paths):
        """Show dialog to assign multiple images to items with full control"""
        if not file_paths or not self.items:
            return

        dialog = ctk.CTkToplevel(self)
        dialog.title(f"批量分配 {len(file_paths)} 张图片")
        dialog.geometry("1100x750")
        dialog.transient(self)

        # Center dialog
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (1100 // 2)
        y = (dialog.winfo_screenheight() // 2) - (750 // 2)
        dialog.geometry(f"1100x750+{x}+{y}")

        # Title and quick assign buttons
        title_frame = ctk.CTkFrame(dialog, fg_color="transparent")
        title_frame.pack(fill="x", padx=20, pady=20)

        ctk.CTkLabel(
            title_frame,
            text=f"为 {len(file_paths)} 张图片选择目标项目",
            font=ctk.CTkFont(size=18, weight="bold")
        ).pack(side="left")

        quick_frame = ctk.CTkFrame(title_frame, fg_color="transparent")
        quick_frame.pack(side="right")

        # Store assignments
        self.assignments = {}

        def quick_assign_first():
            for var in self.assignments.values():
                var.set(1)
            messagebox.showinfo("完成", "已将所有图片分配到第一个项目")

        def quick_assign_even():
            total = len(self.items)
            for i, var in enumerate(self.assignments.values()):
                var.set((i % total) + 1)
            messagebox.showinfo("完成", f"已将图片平均分配到 {total} 个项目")

        def quick_assign_selected():
            if self.selected_item_index is not None:
                idx = self.selected_item_index + 1
                for var in self.assignments.values():
                    var.set(idx)
                messagebox.showinfo("完成", f"已将所有图片分配到项目 {idx}")
            else:
                messagebox.showwarning("提示", "请先选择一个项目")

        ctk.CTkButton(
            quick_frame,
            text="全部→项目1",
            command=quick_assign_first,
            height=32,
            corner_radius=8
        ).pack(side="left", padx=2)

        ctk.CTkButton(
            quick_frame,
            text="平均分配",
            command=quick_assign_even,
            height=32,
            corner_radius=8
        ).pack(side="left", padx=2)

        ctk.CTkButton(
            quick_frame,
            text="→选中项目",
            command=quick_assign_selected,
            height=32,
            corner_radius=8
        ).pack(side="left", padx=2)

        # List frame with headers
        list_container = ctk.CTkFrame(dialog)
        list_container.pack(fill="both", expand=True, padx=20, pady=(0, 10))

        # Headers
        header_frame = ctk.CTkFrame(list_container, fg_color=self.colors['bg_tertiary'], height=40)
        header_frame.pack(fill="x", padx=5, pady=5)
        header_frame.pack_propagate(False)

        headers_info = [
            ("序号", 60),
            ("文件名", 350),
            ("分配到项目", 180),
            ("项目描述", 350)
        ]

        col_x = 10
        for header_text, width in headers_info:
            ctk.CTkLabel(
                header_frame,
                text=header_text,
                font=ctk.CTkFont(size=12, weight="bold"),
                width=width,
                anchor="w"
            ).place(x=col_x, y=10)
            col_x += width + 10

        # Scrollable content
        scroll_frame = ctk.CTkScrollableFrame(list_container, fg_color="transparent")
        scroll_frame.pack(fill="both", expand=True, padx=5, pady=(0, 5))

        # Create project options
        project_options = []
        for j, item in enumerate(self.items):
            desc = item['description'][:25] + ("..." if len(item['description']) > 25 else "")
            project_options.append(f"{j+1}. {desc}")

        # Add row for each file
        for i, fp in enumerate(file_paths):
            row_frame = ctk.CTkFrame(scroll_frame, fg_color=self.colors['bg_secondary'], corner_radius=8)
            row_frame.pack(fill="x", pady=3)

            row_inner = ctk.CTkFrame(row_frame, fg_color="transparent", height=50)
            row_inner.pack(fill="x", padx=10, pady=8)
            row_inner.pack_propagate(False)

            # Index
            ctk.CTkLabel(
                row_inner,
                text=str(i+1),
                width=60,
                font=ctk.CTkFont(size=11)
            ).place(x=0, y=12)

            # Filename
            name = os.path.basename(fp)
            disp = name if len(name) <= 35 else name[:32] + "..."
            ctk.CTkLabel(
                row_inner,
                text=disp,
                width=350,
                anchor="w",
                font=ctk.CTkFont(size=11)
            ).place(x=70, y=12)

            # Project selector
            var = tk.IntVar(value=1)
            combo = tk.ttk.Combobox(
                row_inner,
                width=20,
                state='readonly',
                values=project_options,
                font=('Microsoft YaHei', 10)
            )
            combo.set(project_options[0])
            combo.place(x=430, y=10)

            # Description label
            desc_label = ctk.CTkLabel(
                row_inner,
                text="",
                width=350,
                anchor="w",
                font=ctk.CTkFont(size=10),
                text_color=self.colors['text_secondary']
            )
            desc_label.place(x=620, y=12)

            def update_desc(event=None, v=var, lbl=desc_label, c=combo):
                try:
                    sel = c.get()
                    if sel:
                        pidx = int(sel.split('.')[0]) - 1
                        if 0 <= pidx < len(self.items):
                            v.set(pidx + 1)
                            d = self.items[pidx]['description']
                            lbl.configure(text=d[:30] + ("..." if len(d) > 30 else ""))
                except:
                    pass

            combo.bind('<<ComboboxSelected>>', update_desc)

            # Set initial description
            if self.items:
                first = self.items[0]['description']
                desc_label.configure(text=first[:30] + ("..." if len(first) > 30 else ""))

            self.assignments[fp] = var

        # Bottom buttons
        bottom_frame = ctk.CTkFrame(dialog, fg_color="transparent")
        bottom_frame.pack(fill="x", padx=20, pady=20)

        ctk.CTkLabel(
            bottom_frame,
            text=f"待分配: {len(file_paths)} 张图片 → {len(self.items)} 个项目",
            font=ctk.CTkFont(size=12)
        ).pack(side="left")

        def execute():
            try:
                succ = skip = err = 0
                stats = {}
                for fp, v in self.assignments.items():
                    try:
                        pidx = v.get() - 1
                        if 0 <= pidx < len(self.items):
                            if fp not in self.items[pidx]['images']:
                                self.items[pidx]['images'].append(fp)
                                succ += 1
                                key = f"项目{pidx+1}"
                                stats[key] = stats.get(key, 0) + 1
                            else:
                                skip += 1
                        else:
                            err += 1
                    except:
                        err += 1

                dialog.destroy()
                self.refresh_item_list()
                self.update_stats()

                # Refresh display if item is selected
                if self.selected_item_index is not None and 0 <= self.selected_item_index < len(self.items):
                    self.display_item_images(self.selected_item_index)

                if succ:
                    msg = "批量分配完成！\n\n"
                    msg += f"✅ 成功分配: {succ} 张\n"
                    if skip:
                        msg += f"⚠️ 跳过重复: {skip} 张\n"
                    if err:
                        msg += f"❌ 分配失败: {err} 张\n"
                    if stats:
                        msg += "\n分配详情:\n" + "\n".join([f"  {k}: {v} 张" for k, v in stats.items()])
                    messagebox.showinfo("批量分配完成", msg)
                    self.set_status(f"✅ 批量分配完成：{succ} 张成功")
                else:
                    messagebox.showwarning("分配结果", "没有成功分配任何图片")
                    self.set_status("⚠️ 批量分配：无文件被分配")
            except Exception as e:
                messagebox.showerror("分配错误", f"批量分配失败:\n{str(e)}")
                self.set_status("❌ 批量分配失败")

        ctk.CTkButton(
            bottom_frame,
            text="取消",
            command=dialog.destroy,
            height=40,
            corner_radius=10,
            fg_color=self.colors['bg_tertiary'],
            hover_color=self.colors['border']
        ).pack(side="right", padx=5)

        ctk.CTkButton(
            bottom_frame,
            text="执行分配",
            command=execute,
            height=40,
            corner_radius=10,
            fg_color=self.colors['success'],
            hover_color="#0f6e3d"
        ).pack(side="right", padx=5)

    def display_item_images(self, idx):
        """Display images for selected item"""
        if not (0 <= idx < len(self.items)):
            return

        self.clear_image_display()

        images = self.items[idx].get('images', [])
        if not images:
            # Show drop zone
            return

        # Hide drop zone, show images
        self.drop_zone.pack_forget()

        # Display images in grid
        for img_idx, img_path in enumerate(images):
            row = img_idx // 4
            col = img_idx % 4

            img_card = self.create_image_card(img_path, idx)
            img_card.grid(row=row, column=col, padx=10, pady=10, sticky="nsew")

    def create_image_card(self, img_path, item_idx):
        """Create a modern card for an image"""
        card = ctk.CTkFrame(
            self.image_gallery,
            fg_color=self.colors['bg_secondary'],
            corner_radius=15,
            border_width=2,
            border_color=self.colors['border']
        )

        try:
            # Load and display image
            with Image.open(img_path) as img:
                # Create thumbnail
                img.thumbnail((250, 250), Image.Resampling.LANCZOS)
                photo = ImageTk.PhotoImage(img)

                # Image label
                img_label = ctk.CTkLabel(
                    card,
                    text="",
                    image=photo
                )
                img_label.image = photo  # Keep reference
                img_label.pack(padx=10, pady=10)

            # Image info
            filename = os.path.basename(img_path)
            if len(filename) > 25:
                filename = filename[:22] + "..."

            info_frame = ctk.CTkFrame(card, fg_color=self.colors['bg_tertiary'])
            info_frame.pack(fill="x", padx=10, pady=(0, 10))

            ctk.CTkLabel(
                info_frame,
                text=filename,
                font=ctk.CTkFont(size=11),
                text_color=self.colors['text_secondary']
            ).pack(pady=8)

            # Delete button
            del_btn = ctk.CTkButton(
                card,
                text="删除",
                height=28,
                corner_radius=8,
                fg_color=self.colors['error'],
                hover_color="#C62828",
                command=lambda: self.delete_image(img_path, item_idx)
            )
            del_btn.pack(padx=10, pady=(0, 10))

        except Exception as e:
            ctk.CTkLabel(
                card,
                text=f"无法加载图片\n{str(e)[:30]}",
                text_color=self.colors['error']
            ).pack(pady=20)

        return card

    def delete_image(self, img_path, item_idx):
        """Delete an image from an item"""
        if 0 <= item_idx < len(self.items):
            item = self.items[item_idx]
            if img_path in item['images']:
                if messagebox.askyesno("确认删除", f"确定要删除这张图片吗？\n{os.path.basename(img_path)}"):
                    item['images'].remove(img_path)
                    self.refresh_item_list()
                    self.display_item_images(item_idx)
                    self.update_stats()
                    self.set_status("✓ 已删除图片")

    def clear_image_display(self):
        """Clear image gallery"""
        for widget in self.image_gallery.winfo_children():
            if widget != self.drop_zone:
                widget.destroy()

        # Show drop zone again
        if not self.drop_zone.winfo_viewable():
            self.drop_zone.pack(fill="both", expand=True, padx=20, pady=50)

    def update_stats(self):
        """Update statistics display"""
        total_items = len(self.items)
        total_images = sum(len(item.get('images', [])) for item in self.items)

        self.stats_label.configure(text=f"{total_items} 项目 • {total_images} 图片")

        # Update max images per row
        if self.items:
            self.max_images_per_row = max(len(item.get('images', [])) for item in self.items) or 1

    def set_status(self, message):
        """Update status bar message"""
        self.status_label.configure(text=message)
        self.update_idletasks()

    def _log_debug(self, message):
        """Add a debug log message with timestamp"""
        from datetime import datetime
        timestamp = datetime.now().strftime("%H:%M:%S.%f")[:-3]
        log_entry = f"[{timestamp}] {message}"
        self.debug_logs.append(log_entry)

        # Keep only the most recent logs
        if len(self.debug_logs) > self.max_debug_logs:
            self.debug_logs = self.debug_logs[-self.max_debug_logs:]

        # Also print to console for immediate feedback
        print(log_entry)

    def show_debug_logs(self):
        """Show debug logs window"""
        dialog = ctk.CTkToplevel(self)
        dialog.title("Debug Logs")
        dialog.geometry("900x600")
        dialog.transient(self)

        # Center dialog
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (450)
        y = (dialog.winfo_screenheight() // 2) - (300)
        dialog.geometry(f"900x600+{x}+{y}")

        # Title
        ctk.CTkLabel(
            dialog,
            text="🐛 Debug Logs",
            font=ctk.CTkFont(size=20, weight="bold")
        ).pack(pady=10)

        # Buttons frame
        btn_frame = ctk.CTkFrame(dialog, fg_color="transparent")
        btn_frame.pack(fill="x", padx=20, pady=5)

        def refresh_logs():
            """Refresh the log display"""
            text_widget.configure(state="normal")
            text_widget.delete("1.0", "end")
            for log in self.debug_logs:
                text_widget.insert("end", log + "\n")
            text_widget.configure(state="disabled")
            text_widget.see("end")

        def clear_logs():
            """Clear all logs"""
            self.debug_logs.clear()
            refresh_logs()

        def copy_logs():
            """Copy logs to clipboard"""
            logs_text = "\n".join(self.debug_logs)
            self.clipboard_clear()
            self.clipboard_append(logs_text)
            messagebox.showinfo("已复制", "日志已复制到剪贴板")

        ctk.CTkButton(
            btn_frame,
            text="🔄 刷新",
            command=refresh_logs,
            width=100
        ).pack(side="left", padx=5)

        ctk.CTkButton(
            btn_frame,
            text="🗑️ 清空",
            command=clear_logs,
            width=100
        ).pack(side="left", padx=5)

        ctk.CTkButton(
            btn_frame,
            text="📋 复制",
            command=copy_logs,
            width=100
        ).pack(side="left", padx=5)

        # Log text area with scrollbar
        text_frame = ctk.CTkFrame(dialog)
        text_frame.pack(fill="both", expand=True, padx=20, pady=10)

        text_widget = scrolledtext.ScrolledText(
            text_frame,
            wrap="word",
            font=("Courier", 10),
            bg=self.colors['bg_secondary'],
            fg=self.colors['text_primary'],
            state="disabled"
        )
        text_widget.pack(fill="both", expand=True)

        # Initial load
        refresh_logs()

        # Auto-refresh every 2 seconds
        def auto_refresh():
            if dialog.winfo_exists():
                refresh_logs()
                dialog.after(2000, auto_refresh)

        auto_refresh()

    def show_settings(self):
        """Show settings dialog"""
        dialog = ctk.CTkToplevel(self)
        dialog.title("设置")
        dialog.geometry("500x400")
        dialog.transient(self)

        # Center dialog
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (500 // 2)
        y = (dialog.winfo_screenheight() // 2) - (400 // 2)
        dialog.geometry(f"500x400+{x}+{y}")

        ctk.CTkLabel(
            dialog,
            text="⚙️ 设置",
            font=ctk.CTkFont(size=24, weight="bold")
        ).pack(pady=20)

        # About info
        ctk.CTkLabel(
            dialog,
            text=f"\n系统: {platform.system()}\nPython: {sys.version_info.major}.{sys.version_info.minor}\n拖拽: {'✓' if DRAG_DROP_AVAILABLE else '✗'}\nExcel: {'✓' if EXCEL_AVAILABLE else '✗'}\nPDF: {'✓' if PDF_AVAILABLE else '✗'}",
            font=ctk.CTkFont(size=12),
            text_color=self.colors['text_secondary']
        ).pack(pady=20)

        ctk.CTkButton(
            dialog,
            text="关闭",
            command=dialog.destroy,
            height=40,
            corner_radius=10
        ).pack(pady=20)

    def save_project(self):
        """Save project to JSON file"""
        path = filedialog.asksaveasfilename(
            title="保存项目文件",
            defaultextension=".json",
            filetypes=[("JSON files", "*.json")]
        )

        if path:
            try:
                data = {
                    'title': self.project_title_var.get(),
                    'items': self.items,
                    'created_time': datetime.now().isoformat(),
                    'max_images_per_row': self.max_images_per_row,
                    'version': '2.0.0'
                }

                with open(path, 'w', encoding='utf-8') as f:
                    json.dump(data, f, indent=2, ensure_ascii=False)

                self.set_status(f"✓ 项目已保存: {os.path.basename(path)}")
                messagebox.showinfo("成功", "项目保存成功！")

            except Exception as e:
                messagebox.showerror("错误", f"保存失败: {str(e)}")
                self.set_status("✗ 保存失败")

    def open_project(self):
        """Open project from JSON file"""
        path = filedialog.askopenfilename(
            title="打开项目文件",
            filetypes=[("JSON files", "*.json")]
        )

        if path:
            try:
                with open(path, 'r', encoding='utf-8') as f:
                    data = json.load(f)

                self.project_title_var.set(data.get('title', ''))
                self.items = data.get('items', [])
                self.max_images_per_row = data.get('max_images_per_row', 1)
                self.current_item_id = max((item.get('id', 0) for item in self.items), default=0)

                self.refresh_item_list()
                self.clear_image_display()
                self.update_stats()

                self.set_status(f"✓ 项目已加载: {os.path.basename(path)}")
                messagebox.showinfo("成功", "项目加载成功！")

            except Exception as e:
                messagebox.showerror("错误", f"加载失败: {str(e)}")
                self.set_status("✗ 加载失败")

    def export_excel(self):
        """Export to Excel (reuse original implementation)"""
        if not EXCEL_AVAILABLE:
            messagebox.showerror("错误", "Excel导出功能需要安装openpyxl库\n请运行: pip install openpyxl")
            return

        if not self.items:
            messagebox.showwarning("警告", "没有数据可导出")
            return

        title = self.project_title_var.get().strip()
        if not title:
            if not messagebox.askyesno("标题提醒", "您还没有设置项目标题！\n是否使用默认标题'维修检查报告'继续导出？"):
                return

        path = filedialog.asksaveasfilename(
            title="保存Excel文件",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")]
        )

        if path:
            self.set_status("正在导出Excel...")
            try:
                self._export_excel_file(path)
            except Exception as e:
                messagebox.showerror("错误", f"导出Excel失败: {str(e)}")
                self.set_status("✗ Excel导出失败")

    def _export_excel_file(self, file_path):
        """Export to Excel file (original implementation)"""
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "维修报告"

            title = self.project_title_var.get() or "维修检查报告"

            # Title
            title_cell = ws['A1']
            title_cell.value = title
            title_cell.font = Font(size=20, bold=True, name='微软雅黑')
            title_cell.alignment = Alignment(horizontal='center', vertical='center')

            # Subtitle
            subtitle_cell = ws['A2']
            subtitle_cell.value = f"生成时间：{datetime.now().strftime('%Y年%m月%d日 %H:%M')}"
            subtitle_cell.font = Font(size=11, italic=True, name='微软雅黑')
            subtitle_cell.alignment = Alignment(horizontal='center')

            total_cols = 2 + self.max_images_per_row
            end_col = chr(64 + total_cols)
            ws.merge_cells(f'A1:{end_col}1')
            ws.merge_cells(f'A2:{end_col}2')

            # Headers
            headers = ['序号', '维修内容描述'] + [f'图片{i+1}' for i in range(self.max_images_per_row)]
            for col, header in enumerate(headers, 1):
                c = ws.cell(row=4, column=col)
                c.value = header
                c.font = Font(bold=True, name='微软雅黑')
                c.alignment = Alignment(horizontal='center', vertical='center')
                c.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")

            # Data rows
            temp_files = []
            for row_idx, item in enumerate(self.items, 5):
                ws.cell(row=row_idx, column=1).value = row_idx - 4
                ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal='center', vertical='center')

                desc_cell = ws.cell(row=row_idx, column=2)
                desc_cell.value = item['description']
                desc_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                desc_cell.font = Font(name='微软雅黑')

                images = item.get('images', [])
                row_max_height = 40

                for img_idx, img_path in enumerate(images[:self.max_images_per_row]):
                    col = 3 + img_idx
                    try:
                        if not os.path.exists(img_path):
                            ws.cell(row=row_idx, column=col).value = f"图片文件不存在:\n{os.path.basename(img_path)}"
                            continue

                        with Image.open(img_path) as img:
                            target_w, target_h = 1200, 900
                            r = img.width / img.height
                            if r > target_w/target_h:
                                new_w, new_h = target_w, int(target_w / r)
                            else:
                                new_h, new_w = target_h, int(target_h * r)

                            processed = img.resize((new_w, new_h), Image.Resampling.LANCZOS)
                            tpath = os.path.join(tempfile.gettempdir(), f"excel_img_{uuid.uuid4().hex}.png")
                            processed.save(tpath, 'PNG')

                        temp_files.append(tpath)
                        excel_img = xl_image.Image(tpath)
                        scale = 0.32
                        excel_img.width = new_w * scale
                        excel_img.height = new_h * scale
                        ws.add_image(excel_img, f'{chr(64 + col)}{row_idx}')
                        row_max_height = max(row_max_height, new_h * scale * 0.8)

                    except Exception as e:
                        ws.cell(row=row_idx, column=col).value = f"图片处理失败:\n{os.path.basename(img_path)}"

                ws.row_dimensions[row_idx].height = row_max_height

            # Column widths
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 45
            for i in range(self.max_images_per_row):
                ws.column_dimensions[chr(67+i)].width = 52

            # Borders
            thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
            for row in range(4, len(self.items)+5):
                for col in range(1, total_cols+1):
                    ws.cell(row=row, column=col).border = thin

            wb.save(file_path)

            # Cleanup temp files after delay
            self.after(5000, lambda: self._cleanup_temp_files(temp_files))

            self.set_status(f"✓ Excel文件已保存: {os.path.basename(file_path)}")
            messagebox.showinfo("成功", f"Excel文件已保存到:\n{file_path}")

        except Exception as e:
            messagebox.showerror("错误", f"导出Excel失败: {str(e)}")
            self.set_status("✗ Excel导出失败")

    def export_pdf(self):
        """Export to PDF (reuse original implementation)"""
        if not PDF_AVAILABLE:
            messagebox.showerror("错误", "PDF导出功能需要安装reportlab库\n请运行: pip install reportlab")
            return

        if not self.items:
            messagebox.showwarning("警告", "没有数据可导出")
            return

        title = self.project_title_var.get().strip()
        if not title:
            if not messagebox.askyesno("标题提醒", "您还没有设置项目标题！\n是否使用默认标题'维修检查报告'继续导出？"):
                return

        path = filedialog.asksaveasfilename(
            title="保存PDF文件",
            defaultextension=".pdf",
            filetypes=[("PDF files", "*.pdf")]
        )

        if path:
            self.set_status("正在导出PDF...")
            try:
                self._export_pdf_file(path)
            except Exception as e:
                messagebox.showerror("错误", f"导出PDF失败: {str(e)}")
                self.set_status("✗ PDF导出失败")

    def _export_pdf_file(self, file_path):
        """Export to PDF file (original implementation)"""
        try:
            self._setup_chinese_fonts()

            doc = SimpleDocTemplate(
                file_path,
                pagesize=A4,
                topMargin=20*mm,
                bottomMargin=20*mm,
                leftMargin=15*mm,
                rightMargin=15*mm
            )

            story = []
            styles = getSampleStyleSheet()

            try:
                chinese = ParagraphStyle(
                    'Chinese',
                    parent=styles['Normal'],
                    fontName='Chinese',
                    fontSize=10,
                    leading=12,
                    wordWrap='CJK'
                )
                title_style = ParagraphStyle(
                    'ChineseTitle',
                    parent=styles['Heading1'],
                    fontName='Chinese',
                    fontSize=20,
                    spaceAfter=20,
                    alignment=TA_CENTER,
                    leading=24
                )
                subtitle_style = ParagraphStyle(
                    'Subtitle',
                    parent=chinese,
                    fontSize=11,
                    alignment=TA_CENTER,
                    textColor=colors.HexColor('#666666')
                )
            except:
                chinese = styles['Normal']
                title_style = styles['Heading1']
                subtitle_style = styles['Normal']

            title = self.project_title_var.get() or "维修检查报告"
            story.append(Paragraph(title, title_style))
            story.append(Paragraph(
                f"生成时间：{datetime.now().strftime('%Y年%m月%d日 %H:%M')}",
                subtitle_style
            ))
            story.append(Spacer(1, 30))

            # Add items
            temp_files = []
            for idx, item in enumerate(self.items):
                title_para = Paragraph(
                    f"{idx+1}. {item['description']}",
                    ParagraphStyle(
                        'ItemTitle',
                        parent=chinese,
                        fontSize=12,
                        fontName='Chinese',
                        spaceBefore=10,
                        spaceAfter=10,
                        leftIndent=0,
                        leading=14
                    )
                )
                story.append(title_para)

                images = item.get('images', [])
                if images:
                    img_element = self._create_pdf_images(images, temp_files)
                    if img_element:
                        story.append(img_element)
                    else:
                        story.append(Paragraph("图片加载失败", chinese))
                else:
                    story.append(Paragraph("暂无图片", chinese))

                if idx < len(self.items) - 1:
                    story.append(Spacer(1, 20))

            doc.build(story)

            # Cleanup
            self.after(5000, lambda: self._cleanup_temp_files(temp_files))

            self.set_status(f"✓ PDF文件已保存: {os.path.basename(file_path)}")
            messagebox.showinfo("成功", f"PDF文件已保存到:\n{file_path}")

        except Exception as e:
            messagebox.showerror("错误", f"导出PDF失败: {str(e)}")
            self.set_status("✗ PDF导出失败")

    def _setup_chinese_fonts(self):
        """Setup Chinese fonts for PDF"""
        try:
            system = platform.system()
            if system == "Windows":
                paths = ["C:/Windows/Fonts/simsun.ttc", "C:/Windows/Fonts/simhei.ttf", "C:/Windows/Fonts/msyh.ttc"]
            elif system == "Darwin":
                paths = ["/Library/Fonts/Arial Unicode.ttf", "/System/Library/Fonts/PingFang.ttc"]
            else:
                paths = ["/usr/share/fonts/truetype/wqy/wqy-microhei.ttc", "/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc"]

            for p in paths:
                if os.path.exists(p):
                    pdfmetrics.registerFont(TTFont('Chinese', p))
                    return

            pdfmetrics.registerFont(TTFont('Chinese', 'Helvetica'))
        except:
            pass

    def _create_pdf_images(self, images, temp_files):
        """Create PDF image layout"""
        try:
            if len(images) == 1:
                p = images[0]
                if os.path.exists(p):
                    t = self._process_pdf_image(p, temp_files, 150*mm, 100*mm)
                    if t:
                        img = RL_Image(t, width=150*mm, height=100*mm, kind='proportional')
                        table = Table([[img]], colWidths=[170*mm])
                        table.setStyle(TableStyle([('ALIGN', (0,0), (-1,-1), 'CENTER')]))
                        return table
            else:
                cols = 2 if len(images) <= 4 else 3
                rows, row = [], []
                for i, p in enumerate(images):
                    if os.path.exists(p):
                        size = 70*mm if cols==2 else 50*mm
                        t = self._process_pdf_image(p, temp_files, size, size)
                        row.append(RL_Image(t, width=size, height=size, kind='proportional') if t else "")
                    else:
                        row.append("")

                    if len(row) >= cols or i == len(images)-1:
                        while len(row) < cols:
                            row.append("")
                        rows.append(row)
                        row = []

                if rows:
                    col_w = 85*mm if cols==2 else 56*mm
                    table = Table(rows, colWidths=[col_w]*cols)
                    table.setStyle(TableStyle([
                        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                        ('LEFTPADDING', (0,0), (-1,-1), 5),
                        ('RIGHTPADDING', (0,0), (-1,-1), 5),
                        ('TOPPADDING', (0,0), (-1,-1), 5),
                        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
                    ]))
                    return table
        except:
            pass
        return None

    def _process_pdf_image(self, img_path, temp_files, max_width, max_height):
        """Process image for PDF"""
        try:
            with Image.open(img_path) as img:
                max_w_px, max_h_px = int(max_width*10), int(max_height*10)
                r = img.width / img.height
                if img.width/max_w_px > img.height/max_h_px:
                    new_w, new_h = max_w_px, int(max_w_px / r)
                else:
                    new_h, new_w = max_h_px, int(max_h_px * r)

                resized = img.resize((new_w, new_h), Image.Resampling.LANCZOS)
                if resized.mode != 'RGB':
                    resized = resized.convert('RGB')

                t = os.path.join(tempfile.gettempdir(), f"pdf_img_{uuid.uuid4().hex}.jpg")
                resized.save(t, 'JPEG', quality=92)
                temp_files.append(t)
                return t
        except:
            return None

    def _cleanup_temp_files(self, files):
        """Cleanup temporary files"""
        for f in files:
            try:
                if os.path.exists(f):
                    os.unlink(f)
            except:
                pass

    def create_menu(self):
        """Create menu bar with all options"""
        menubar = tk.Menu(self)
        self.config(menu=menubar)

        # File menu
        file_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="文件", menu=file_menu)
        file_menu.add_command(label="新建项目", command=self.new_project, accelerator="Ctrl+N")
        file_menu.add_command(label="打开项目", command=self.open_project, accelerator="Ctrl+O")
        file_menu.add_command(label="保存项目", command=self.save_project, accelerator="Ctrl+S")
        file_menu.add_separator()
        if EXCEL_AVAILABLE:
            file_menu.add_command(label="导出Excel", command=self.export_excel, accelerator="Ctrl+E")
        if PDF_AVAILABLE:
            file_menu.add_command(label="导出PDF", command=self.export_pdf, accelerator="Ctrl+P")
        file_menu.add_separator()
        file_menu.add_command(label="退出", command=self.quit)

        # Edit menu
        edit_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="编辑", menu=edit_menu)
        edit_menu.add_command(label="添加项目", command=self.add_item, accelerator="Ctrl+A")
        edit_menu.add_command(label="删除项目", command=self.delete_selected_item, accelerator="Delete")
        edit_menu.add_separator()
        edit_menu.add_command(label="批量添加图片", command=self.batch_add_images, accelerator="Ctrl+I")

        # View menu
        view_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="视图", menu=view_menu)
        view_menu.add_command(label="预览报告", command=self.preview_report)
        view_menu.add_command(label="刷新", command=self.refresh_display, accelerator="F5")

        # Help menu
        help_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="帮助", menu=help_menu)
        help_menu.add_command(label="使用说明", command=self.show_help)
        help_menu.add_command(label="关于", command=self.show_about)
        help_menu.add_separator()
        help_menu.add_command(label="Debug Logs", command=self.show_debug_logs, accelerator="Ctrl+D")

    def bind_shortcuts(self):
        """Bind keyboard shortcuts"""
        self.bind('<Control-n>', lambda e: self.new_project())
        self.bind('<Control-o>', lambda e: self.open_project())
        self.bind('<Control-s>', lambda e: self.save_project())
        self.bind('<Control-e>', lambda e: self.export_excel())
        self.bind('<Control-p>', lambda e: self.export_pdf())
        self.bind('<Control-a>', lambda e: self.add_item())
        self.bind('<Control-i>', lambda e: self.batch_add_images())
        self.bind('<Control-d>', lambda e: self.show_debug_logs())
        self.bind('<Delete>', lambda e: self.delete_selected_item())
        self.bind('<F5>', lambda e: self.refresh_display())

    def new_project(self):
        """Create new project"""
        if self.items and messagebox.askyesno("确认", "当前项目未保存，确定要新建项目吗？"):
            self.items = []
            self.current_item_id = 0
            self.project_title_var.set("")
            self.selected_item_index = None
            self.refresh_item_list()
            self.clear_image_display()
            self.description_entry.delete(0, "end")
            self.thumbnail_cache.clear()
            self.update_stats()
            self.set_status("✓ 已创建新项目")
        elif not self.items:
            self.project_title_var.set("")
            self.set_status("新项目就绪")

    def delete_selected_item(self):
        """Delete currently selected item"""
        if self.selected_item_index is not None:
            self.delete_item(self.selected_item_index)
        else:
            messagebox.showwarning("警告", "请先选择要删除的项目")

    def refresh_display(self):
        """Refresh the display"""
        self.refresh_item_list()
        if self.selected_item_index is not None and 0 <= self.selected_item_index < len(self.items):
            self.display_item_images(self.selected_item_index)
        self.update_stats()
        self.set_status("✓ 已刷新显示")

    def preview_report(self):
        """Preview report content"""
        if not self.items:
            messagebox.showwarning("警告", "没有数据可预览")
            return

        dialog = ctk.CTkToplevel(self)
        dialog.title("报告预览")
        dialog.geometry("1000x700")
        dialog.transient(self)

        # Center dialog
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (1000 // 2)
        y = (dialog.winfo_screenheight() // 2) - (700 // 2)
        dialog.geometry(f"1000x700+{x}+{y}")

        # Create text widget
        text_frame = ctk.CTkFrame(dialog)
        text_frame.pack(fill="both", expand=True, padx=20, pady=20)

        text = tk.Text(text_frame, wrap="word", font=('Courier', 10), bg='white')
        scrollbar = ctk.CTkScrollbar(text_frame, command=text.yview)
        text.configure(yscrollcommand=scrollbar.set)

        scrollbar.pack(side="right", fill="y")
        text.pack(side="left", fill="both", expand=True)

        # Generate content
        title = self.project_title_var.get() or "维修检查报告"
        content = f"{'='*60}\n{title:^60}\n{'='*60}\n\n"
        content += f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
        content += f"项目总数: {len(self.items)}\n"
        content += f"图片总数: {sum(len(it.get('images',[])) for it in self.items)}\n"
        content += f"工具版本: v2.0.0 Modern Edition\n\n"

        for i, item in enumerate(self.items):
            content += f"{'-'*60}\n项目 {i+1}: {item['description']}\n{'-'*60}\n"
            imgs = item.get('images', [])
            if imgs:
                content += f"包含图片 ({len(imgs)} 张):\n"
                for j, p in enumerate(imgs):
                    try:
                        size_k = os.path.getsize(p)/1024
                        with Image.open(p) as im:
                            info = f"{im.width}×{im.height}"
                        content += f"  {j+1}. {os.path.basename(p)} ({size_k:.1f}KB, {info})\n"
                    except:
                        content += f"  {j+1}. {os.path.basename(p)} (无法读取信息)\n"
            else:
                content += "暂无图片\n"
            content += "\n"

        text.insert("1.0", content)
        text.config(state="disabled")

    def show_help(self):
        """Show help dialog"""
        dialog = ctk.CTkToplevel(self)
        dialog.title("使用说明")
        dialog.geometry("800x600")
        dialog.transient(self)

        # Center dialog
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (800 // 2)
        y = (dialog.winfo_screenheight() // 2) - (600 // 2)
        dialog.geometry(f"800x600+{x}+{y}")

        # Title
        ctk.CTkLabel(
            dialog,
            text="📖 使用说明",
            font=ctk.CTkFont(size=24, weight="bold")
        ).pack(pady=20)

        # Content frame
        content_frame = ctk.CTkFrame(dialog)
        content_frame.pack(fill="both", expand=True, padx=20, pady=(0, 20))

        help_text = """
维修单工具 Modern Edition v2.0 - 使用说明

🎨 界面说明：
• 左侧面板：显示所有维修项目列表
• 右侧区域：显示当前选中项目的图片
• 顶部工具栏：项目标题输入和导出按钮

📝 基本操作：
1. 创建项目：点击左侧"+ 添加"按钮或使用 Ctrl+A
2. 编辑描述：选中项目后，在右侧描述框中输入
3. 添加图片：选中项目后，点击"添加图片"或"批量添加"
4. 删除项目/图片：点击相应的删除按钮

💾 保存和导出：
• 保存项目：文件 → 保存项目 (Ctrl+S)
• 打开项目：文件 → 打开项目 (Ctrl+O)
• 导出Excel：文件 → 导出Excel (Ctrl+E)
• 导出PDF：文件 → 导出PDF (Ctrl+P)

⌨️ 快捷键：
• Ctrl+N：新建项目
• Ctrl+O：打开项目
• Ctrl+S：保存项目
• Ctrl+A：添加项目
• Ctrl+I：批量添加图片
• Ctrl+E：导出Excel
• Ctrl+P：导出PDF
• Delete：删除选中项目
• F5：刷新显示

📸 图片支持：
支持 JPG, PNG, GIF, BMP, TIFF, WEBP 等格式

💡 提示：
• 导出前请务必填写项目标题
• 建议定期保存项目文件
• 图片会自动优化以适应导出格式
        """

        text_widget = tk.Text(
            content_frame,
            wrap="word",
            font=('Microsoft YaHei', 10),
            bg='white',
            padx=20,
            pady=20
        )
        scrollbar = ctk.CTkScrollbar(content_frame, command=text_widget.yview)
        text_widget.configure(yscrollcommand=scrollbar.set)

        scrollbar.pack(side="right", fill="y")
        text_widget.pack(side="left", fill="both", expand=True)

        text_widget.insert("1.0", help_text)
        text_widget.config(state="disabled")

        # Close button
        ctk.CTkButton(
            dialog,
            text="关闭",
            command=dialog.destroy,
            height=40,
            corner_radius=10
        ).pack(pady=(0, 20))

    def show_about(self):
        """Show about dialog"""
        about_text = f"""维修单工具 Modern Edition v2.0

🎨 明亮透明主题设计
📊 支持Excel/PDF导出
🖼️ 智能图片处理

系统：{platform.system()} {platform.release()}
Python：{sys.version_info.major}.{sys.version_info.minor}.{sys.version_info.micro}

功能状态：
拖拽：{'✅ 可用' if DRAG_DROP_AVAILABLE else '❌ 不可用'}
Excel：{'✅ 可用' if EXCEL_AVAILABLE else '❌ 不可用'}
PDF：{'✅ 可用' if PDF_AVAILABLE else '❌ 不可用'}
"""
        messagebox.showinfo("关于", about_text)

    def quick_fill_title(self):
        """Quick fill title with templates"""
        dialog = ctk.CTkToplevel(self)
        dialog.title("选择标题模板")
        dialog.geometry("500x400")
        dialog.transient(self)

        # Center dialog
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (500 // 2)
        y = (dialog.winfo_screenheight() // 2) - (400 // 2)
        dialog.geometry(f"500x400+{x}+{y}")

        # Title
        ctk.CTkLabel(
            dialog,
            text="选择或编辑标题模板",
            font=ctk.CTkFont(size=18, weight="bold")
        ).pack(pady=20)

        # Template list
        current_time = datetime.now()
        suggestions = [
            f"{current_time.strftime('%Y年%m月')} 设备维修检查报告",
            f"{current_time.strftime('%Y-%m-%d')} 维修作业报告",
            "设备保养维护记录",
            "故障排查维修报告",
            "定期检修报告"
        ]

        list_frame = ctk.CTkFrame(dialog)
        list_frame.pack(fill="both", expand=True, padx=20, pady=(0, 10))

        listbox = tk.Listbox(
            list_frame,
            font=('Microsoft YaHei', 11),
            bg='white',
            selectmode=tk.SINGLE,
            activestyle='dotbox'
        )
        listbox.pack(fill="both", expand=True, padx=5, pady=5)

        for s in suggestions:
            listbox.insert(tk.END, s)
        listbox.selection_set(0)

        # Custom entry
        custom_frame = ctk.CTkFrame(dialog)
        custom_frame.pack(fill="x", padx=20, pady=10)

        ctk.CTkLabel(
            custom_frame,
            text="或自定义：",
            font=ctk.CTkFont(size=12)
        ).pack(anchor="w", pady=(0, 5))

        custom_entry = ctk.CTkEntry(
            custom_frame,
            placeholder_text="输入自定义标题...",
            height=36,
            corner_radius=8
        )
        custom_entry.pack(fill="x")

        # Buttons
        btn_frame = ctk.CTkFrame(dialog, fg_color="transparent")
        btn_frame.pack(fill="x", padx=20, pady=(0, 20))

        def apply_title():
            sel = listbox.curselection()
            if sel:
                self.project_title_var.set(suggestions[sel[0]])
            elif custom_entry.get().strip():
                self.project_title_var.set(custom_entry.get().strip())
            dialog.destroy()
            self.set_status("✓ 标题已应用")

        ctk.CTkButton(
            btn_frame,
            text="取消",
            command=dialog.destroy,
            height=36,
            corner_radius=8,
            fg_color=self.colors['bg_tertiary'],
            hover_color=self.colors['border']
        ).pack(side="right", padx=5)

        ctk.CTkButton(
            btn_frame,
            text="确定",
            command=apply_title,
            height=36,
            corner_radius=8,
            fg_color=self.colors['accent'],
            hover_color=self.colors['accent_hover']
        ).pack(side="right")

        # Bind double-click and Enter
        listbox.bind('<Double-Button-1>', lambda e: apply_title())
        dialog.bind('<Return>', lambda e: apply_title())
        dialog.bind('<Escape>', lambda e: dialog.destroy())

    def move_item_up(self):
        """Move selected item up"""
        if self.selected_item_index is None:
            messagebox.showwarning("提示", "请先选择一个项目")
            return

        idx = self.selected_item_index
        if idx > 0:
            self.items[idx], self.items[idx-1] = self.items[idx-1], self.items[idx]
            self.selected_item_index = idx - 1
            self.refresh_item_list()
            self.set_status("✓ 项目已上移")
        else:
            self.set_status("已经是第一个项目")

    def move_item_down(self):
        """Move selected item down"""
        if self.selected_item_index is None:
            messagebox.showwarning("提示", "请先选择一个项目")
            return

        idx = self.selected_item_index
        if idx < len(self.items) - 1:
            self.items[idx], self.items[idx+1] = self.items[idx+1], self.items[idx]
            self.selected_item_index = idx + 1
            self.refresh_item_list()
            self.set_status("✓ 项目已下移")
        else:
            self.set_status("已经是最后一个项目")


def main():
    """Main entry point"""
    print("="*60)
    print("🚀 启动维修单工具 Modern Edition v2.0")
    print("="*60)
    print(f"系统: {platform.system()}")
    print(f"Python: {sys.version_info.major}.{sys.version_info.minor}")
    print(f"拖拽: {'✓' if DRAG_DROP_AVAILABLE else '✗'}")
    print(f"Excel: {'✓' if EXCEL_AVAILABLE else '✗'}")
    print(f"PDF: {'✓' if PDF_AVAILABLE else '✗'}")
    print("="*60)

    # Check for CustomTkinter
    try:
        import customtkinter
        print("✓ CustomTkinter 已加载")
    except ImportError:
        print("✗ CustomTkinter 未安装")
        print("\n请安装 CustomTkinter:")
        print("  pip install customtkinter")
        return

    # Launch app
    try:
        app = ModernRepairTool()
        app.mainloop()
    except Exception as e:
        print(f"❌ 程序启动失败: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()
