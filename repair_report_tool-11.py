#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
维修单工具 - 图文报告生成器 v1.7.4
修复：
-  启动缺少 refresh_display 导致崩溃
-  TkDND 自动探测新增 Homebrew opt 路径；更稳健的日志与回退
-  保留：标题输入法提交修复、拖拽路径解析
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
from PIL import Image, ImageTk, ImageFilter
import os
import json
from datetime import datetime
from pathlib import Path
import tempfile
import uuid
import platform
import atexit
import re
import sys
import glob

# Excel导出
try:
    import openpyxl
    from openpyxl.drawing import image as xl_image
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# PDF导出
try:
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RL_Image, PageBreak, KeepTogether
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch, mm
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

# 拖拽功能
DRAG_DROP_AVAILABLE = False
try:
    from tkinterdnd2 import DND_FILES, TkinterDnD
    import tkinterdnd2 as _tkdnd_pkg
    DRAG_DROP_AVAILABLE = True
except ImportError:
    _tkdnd_pkg = None

class RepairReportTool:
    def __init__(self):
        self.drag_drop_working = False
        self.root = self._init_root_with_dnd()

        self.root.title("维修单工具 - 图文报告生成器 v1.7.4")
        self.root.geometry("1400x900")
        try:
            self.root.state('zoomed')
        except:
            pass

        self.project_title = tk.StringVar(value="")
        self.items = []
        self.current_item_id = 0
        self.max_images_per_row = 1

        self.image_cache = {}
        self.thumbnail_cache = {}

        self.right_frame = None
        self.img_display_frame = None
        self.selected_images = set()

        self.setup_ui()
        self.update_max_images_per_row()

    # ---------- TkDND 初始化与自动探测 ----------
    def _init_root_with_dnd(self):
        if not DRAG_DROP_AVAILABLE:
            print("⚠️ 未安装 tkinterdnd2，拖拽不可用")
            self.drag_drop_working = False
            return tk.Tk()
        # 尝试直接初始化
        try:
            root = TkinterDnD.Tk()
            self.drag_drop_working = True
            print("✅ 拖拽功能初始化成功")
            return root
        except Exception as e:
            print(f"⚠️ 拖拽功能初始化失败: {e}")

        # 二次尝试：自动定位 TkDND2.9 并设置 TKDND_LIBRARY
        candidate_dirs = self._collect_tkdnd_candidate_dirs()
        chosen = None
        for d in candidate_dirs:
            if d and os.path.isdir(d) and os.path.exists(os.path.join(d, "pkgIndex.tcl")):
                chosen = d
                break
        if chosen:
            os.environ["TKDND_LIBRARY"] = chosen
            print(f"🔧 设定 TKDND_LIBRARY={chosen} 后再次尝试初始化...")
            try:
                root = TkinterDnD.Tk()
                self.drag_drop_working = True
                print("✅ 拖拽功能初始化成功（通过 TKDND_LIBRARY）")
                return root
            except Exception as e2:
                print(f"❌ 二次初始化失败: {e2}")

        # 兜底
        self.drag_drop_working = False
        print("ℹ️ 已回退到普通 Tk，拖拽不可用")
        return tk.Tk()

    def _collect_tkdnd_candidate_dirs(self):
        candidates = []
        # 环境变量
        env_path = os.environ.get("TKDND_LIBRARY")
        if env_path:
            candidates.append(env_path)
        # 包内可能路径
        try:
            if _tkdnd_pkg:
                pkg_dir = Path(_tkdnd_pkg.__file__).parent
                for name in ["TkDND2.9", "tkdnd2.9", "tkdnd", "TKDND2.9"]:
                    candidates.append(str(pkg_dir / name))
        except Exception:
            pass
        # Homebrew Cellar 与 opt（新增 opt 稳定链接）
        cellar_bases = ["/opt/homebrew/Cellar/tkdnd", "/usr/local/Cellar/tkdnd"]
        for base in cellar_bases:
            for p in glob.glob(os.path.join(base, "*/lib/TkDND2.9")):
                candidates.append(p)
        opt_bases = ["/opt/homebrew/opt/tkdnd/lib/TkDND2.9", "/usr/local/opt/tkdnd/lib/TkDND2.9"]
        candidates += opt_bases
        # 系统常见目录
        candidates += [
            "/Library/Tcl/TkDND2.9",
            "/Library/Frameworks/Tk.framework/Versions/8.6/Resources/Scripts/TkDND2.9",
            "/usr/local/lib/TkDND2.9",
            "/usr/lib/TkDND2.9",
        ]
        # Python 前缀
        py_prefix = Path(sys.prefix)
        candidates += [
            str(py_prefix / "lib" / "TkDND2.9"),
            str(py_prefix / "Library" / "Tcl" / "TkDND2.9"),
        ]
        # 去重
        seen, uniq = set(), []
        for c in candidates:
            if c and c not in seen:
                seen.add(c); uniq.append(c)
        return uniq
    # ---------- 结束：TkDND 初始化 ----------

    def setup_ui(self):
        self.create_menu()
        self.create_toolbar()
        self.create_main_content()
        self.create_statusbar()

    def create_menu(self):
        menubar = tk.Menu(self.root)
        self.root.config(menu=menubar)

        file_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="文件", menu=file_menu)
        file_menu.add_command(label="新建项目", command=self.new_project, accelerator="Ctrl+N")
        file_menu.add_command(label="打开项目", command=self.open_project, accelerator="Ctrl+O")
        file_menu.add_command(label="保存项目", command=self.save_project, accelerator="Ctrl+S")
        file_menu.add_separator()
        file_menu.add_command(label="导出Excel", command=self.export_excel, accelerator="Ctrl+E")
        file_menu.add_command(label="导出PDF", command=self.export_pdf, accelerator="Ctrl+P")
        file_menu.add_separator()
        file_menu.add_command(label="退出", command=self.root.quit)

        edit_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="编辑", menu=edit_menu)
        edit_menu.add_command(label="添加项目", command=self.add_item, accelerator="Ctrl+A")
        edit_menu.add_command(label="删除项目", command=self.delete_selected_item, accelerator="Delete")
        edit_menu.add_separator()
        edit_menu.add_command(label="批量添加图片", command=self.batch_add_images, accelerator="Ctrl+I")

        view_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="视图", menu=view_menu)
        view_menu.add_command(label="预览报告", command=self.preview_report)
        view_menu.add_command(label="刷新", command=self.refresh_display, accelerator="F5")

        help_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="帮助", menu=help_menu)
        help_menu.add_command(label="使用说明", command=self.show_help)
        help_menu.add_command(label="关于", command=self.show_about)

        self.root.bind('<Control-n>', lambda e: self.new_project())
        self.root.bind('<Control-o>', lambda e: self.open_project())
        self.root.bind('<Control-s>', lambda e: self.save_project())
        self.root.bind('<Control-e>', lambda e: self.export_excel())
        self.root.bind('<Control-p>', lambda e: self.export_pdf())
        self.root.bind('<Control-a>', lambda e: self.add_item())
        self.root.bind('<Control-i>', lambda e: self.batch_add_images())
        self.root.bind('<Delete>', lambda e: self.delete_selected_item())
        self.root.bind('<F5>', lambda e: self.refresh_display())

    def create_toolbar(self):
        toolbar_frame = ttk.Frame(self.root)
        toolbar_frame.pack(side=tk.TOP, fill=tk.X, padx=5, pady=5)

        title_frame = ttk.LabelFrame(toolbar_frame, text="📋 项目标题（必填！导出文档的标题）", padding="8")
        title_frame.pack(side=tk.LEFT, padx=5)

        title_entry_frame = ttk.Frame(title_frame)
        title_entry_frame.pack(fill=tk.X)

        self.title_entry = ttk.Entry(title_entry_frame, textvariable=self.project_title,
                                     width=40, font=('Arial', 11, 'bold'))
        self.title_entry.pack(side=tk.LEFT, padx=5)

        self.title_status_label = ttk.Label(title_frame, text="⚠️ 请输入项目标题",
                                            font=('Arial', 9), foreground='orange')
        self.title_status_label.pack()

        def validate_title(*_):
            title = self.project_title.get().strip()
            if title:
                self.title_status_label.config(text=f"✅ 当前标题：{title[:30]}{'...' if len(title)>30 else ''}",
                                               foreground='green')
            else:
                self.title_status_label.config(text="⚠️ 请输入项目标题", foreground='orange')

        try:
            self.project_title.trace_add('write', validate_title)
        except AttributeError:
            try:
                self.project_title.trace('w', validate_title)
            except:
                self.title_entry.bind('<KeyRelease>', validate_title)
                self.title_entry.bind('<FocusOut>', validate_title)

        ttk.Button(title_entry_frame, text="快速填充",
                   command=self.quick_fill_title, width=8).pack(side=tk.LEFT, padx=2)
        ttk.Button(title_entry_frame, text="应用标题",
                   command=lambda: (self.get_project_title(), self.set_status("标题已应用")),
                   width=8).pack(side=tk.LEFT, padx=2)

        validate_title()

        ttk.Separator(toolbar_frame, orient=tk.VERTICAL).pack(side=tk.LEFT, fill=tk.Y, padx=10)

        ttk.Button(toolbar_frame, text="新建", command=self.new_project).pack(side=tk.LEFT, padx=2)
        ttk.Button(toolbar_frame, text="打开", command=self.open_project).pack(side=tk.LEFT, padx=2)
        ttk.Button(toolbar_frame, text="保存", command=self.save_project).pack(side=tk.LEFT, padx=2)

        ttk.Separator(toolbar_frame, orient=tk.VERTICAL).pack(side=tk.LEFT, fill=tk.Y, padx=10)

        ttk.Button(toolbar_frame, text="添加项目", command=self.add_item).pack(side=tk.LEFT, padx=2)
        ttk.Button(toolbar_frame, text="批量图片", command=self.batch_add_images).pack(side=tk.LEFT, padx=2)
        ttk.Button(toolbar_frame, text="预览", command=self.preview_report).pack(side=tk.LEFT, padx=2)

        ttk.Separator(toolbar_frame, orient=tk.VERTICAL).pack(side=tk.LEFT, fill=tk.Y, padx=10)

        export_frame = ttk.Frame(toolbar_frame)
        export_frame.pack(side=tk.RIGHT, padx=5)
        if EXCEL_AVAILABLE:
            ttk.Button(export_frame, text="导出Excel", command=self.export_excel).pack(side=tk.RIGHT, padx=2)
        if PDF_AVAILABLE:
            ttk.Button(export_frame, text="导出PDF", command=self.export_pdf).pack(side=tk.RIGHT, padx=2)

        self.stats_label = ttk.Label(toolbar_frame, text="项目: 0 | 图片: 0")
        self.stats_label.pack(side=tk.RIGHT, padx=10)

    def quick_fill_title(self):
        current_time = datetime.now()
        suggestions = [
            f"{current_time.strftime('%Y年%m月')} 设备维修检查报告",
            f"{current_time.strftime('%Y-%m-%d')} 维修作业报告",
            "设备保养维护记录",
            "故障排查维修报告",
            "定期检修报告"
        ]
        dialog = tk.Toplevel(self.root)
        dialog.title("选择标题模板")
        dialog.geometry("450x350")
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.geometry("+%d+%d" % (self.root.winfo_rootx()+50, self.root.winfo_rooty()+50))
        ttk.Label(dialog, text="选择或编辑标题模板：", font=('Arial', 11, 'bold')).pack(pady=10)
        listbox = tk.Listbox(dialog, font=('Arial', 10), height=6)
        for s in suggestions:
            listbox.insert(tk.END, s)
        listbox.pack(fill=tk.BOTH, expand=True, padx=15, pady=5)
        custom_frame = ttk.Frame(dialog); custom_frame.pack(fill=tk.X, padx=15, pady=10)
        ttk.Label(custom_frame, text="或自定义：", font=('Arial', 10)).pack(anchor=tk.W)
        custom_entry = ttk.Entry(custom_frame, width=50, font=('Arial', 10)); custom_entry.pack(fill=tk.X, pady=(5,0))
        btn_frame = ttk.Frame(dialog); btn_frame.pack(fill=tk.X, padx=15, pady=10)

        def apply_title():
            sel = listbox.curselection()
            if sel:
                self.project_title.set(suggestions[sel[0]])
            elif custom_entry.get().strip():
                self.project_title.set(custom_entry.get().strip())
            dialog.destroy()

        ttk.Button(btn_frame, text="取消", command=dialog.destroy).pack(side=tk.RIGHT, padx=5)
        ttk.Button(btn_frame, text="确定", command=apply_title).pack(side=tk.RIGHT)
        listbox.selection_set(0); listbox.focus_set()
        listbox.bind('<Double-Button-1>', lambda e: apply_title())
        dialog.bind('<Return>', lambda e: apply_title())
        dialog.bind('<Escape>', lambda e: dialog.destroy())

    def create_main_content(self):
        main_paned = ttk.PanedWindow(self.root, orient=tk.HORIZONTAL)
        main_paned.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        self.create_left_panel(main_paned)
        self.create_right_panel(main_paned)

    def create_left_panel(self, parent):
        left_frame = ttk.Frame(parent, width=400)
        parent.add(left_frame, weight=1)
        list_title_frame = ttk.Frame(left_frame); list_title_frame.pack(fill=tk.X, padx=5, pady=5)
        ttk.Label(list_title_frame, text="维修项目列表", font=('Arial', 12, 'bold')).pack(side=tk.LEFT)
        btn_frame = ttk.Frame(list_title_frame); btn_frame.pack(side=tk.RIGHT)
        ttk.Button(btn_frame, text="+", width=3, command=self.add_item).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_frame, text="-", width=3, command=self.delete_selected_item).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_frame, text="↑", width=3, command=self.move_item_up).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_frame, text="↓", width=3, command=self.move_item_down).pack(side=tk.LEFT, padx=2)

        list_frame = ttk.Frame(left_frame); list_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        columns = ('序号','描述','图片数量')
        self.item_tree = ttk.Treeview(list_frame, columns=columns, show='headings', height=15)
        self.item_tree.heading('序号', text='序号')
        self.item_tree.heading('描述', text='维修内容描述')
        self.item_tree.heading('图片数量', text='图片')
        self.item_tree.column('序号', width=50, anchor=tk.CENTER)
        self.item_tree.column('描述', width=250, anchor=tk.W)
        self.item_tree.column('图片数量', width=60, anchor=tk.CENTER)
        tree_scroll = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=self.item_tree.yview)
        self.item_tree.configure(yscrollcommand=tree_scroll.set)
        self.item_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        tree_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.item_tree.bind('<<TreeviewSelect>>', self.on_item_select)
        self.item_tree.bind('<Double-1>', self.edit_item_description)

        desc_frame = ttk.LabelFrame(left_frame, text="项目描述", padding="5")
        desc_frame.pack(fill=tk.X, padx=5, pady=5)
        self.description_text = scrolledtext.ScrolledText(desc_frame, height=5, wrap=tk.WORD)
        self.description_text.pack(fill=tk.BOTH, expand=True)
        self.description_text.bind('<KeyRelease>', self.on_description_change)

    def create_right_panel(self, parent):
        self.right_frame = ttk.Frame(parent)
        parent.add(self.right_frame, weight=2)

        img_title_frame = ttk.Frame(self.right_frame)
        img_title_frame.pack(fill=tk.X, padx=5, pady=5)
        ttk.Label(img_title_frame, text="图片管理", font=('Arial', 12, 'bold')).pack(side=tk.LEFT)

        if self.drag_drop_working:
            drag_label = ttk.Label(img_title_frame, text="📎 支持拖拽图片",
                                   font=('Arial', 9), foreground='green', cursor="arrow")
        else:
            drag_label = ttk.Label(img_title_frame, text="⚠️ 拖拽不可用（点击查看修复）",
                                   font=('Arial', 9), foreground='orange', cursor="hand2")
            drag_label.bind("<Button-1>", lambda e: self.show_dnd_fix_guide())
        drag_label.pack(side=tk.LEFT, padx=20)

        img_btn_frame = ttk.Frame(img_title_frame); img_btn_frame.pack(side=tk.RIGHT)
        ttk.Button(img_btn_frame, text="添加图片", command=self.add_images).pack(side=tk.LEFT, padx=2)
        ttk.Button(img_btn_frame, text="批量添加", command=self.batch_add_images).pack(side=tk.LEFT, padx=2)
        ttk.Button(img_btn_frame, text="删除图片", command=self.delete_selected_images).pack(side=tk.LEFT, padx=2)
        ttk.Button(img_btn_frame, text="清空图片", command=self.clear_images).pack(side=tk.LEFT, padx=2)

        self.img_display_frame = ttk.Frame(self.right_frame)
        self.img_display_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        canvas_frame = ttk.Frame(self.img_display_frame); canvas_frame.pack(fill=tk.BOTH, expand=True)
        self.img_canvas = tk.Canvas(canvas_frame, bg='white')
        img_h_scroll = ttk.Scrollbar(canvas_frame, orient=tk.HORIZONTAL, command=self.img_canvas.xview)
        img_v_scroll = ttk.Scrollbar(canvas_frame, orient=tk.VERTICAL, command=self.img_canvas.yview)
        self.img_canvas.configure(xscrollcommand=img_h_scroll.set, yscrollcommand=img_v_scroll.set)
        img_h_scroll.pack(side=tk.BOTTOM, fill=tk.X)
        img_v_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.img_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        self.img_scroll_frame = ttk.Frame(self.img_canvas)
        self.img_canvas.create_window((0, 0), window=self.img_scroll_frame, anchor="nw")
        self.img_canvas.bind('<MouseWheel>', self._on_mousewheel)
        self.img_canvas.bind('<Button-4>', self._on_mousewheel)
        self.img_canvas.bind('<Button-5>', self._on_mousewheel)
        self.img_scroll_frame.bind('<Configure>', self._on_frame_configure)

        self.setup_drag_and_drop()

    def setup_drag_and_drop(self):
        if not self.drag_drop_working:
            return
        try:
            areas = [self.root, self.img_canvas, self.img_display_frame, self.img_scroll_frame, self.right_frame]
            for area in areas:
                if area and hasattr(area, 'drop_target_register'):
                    area.drop_target_register(DND_FILES)
                    area.dnd_bind('<<Drop>>', self.on_drop)
            print("✅ 拖拽功能设置成功")
        except Exception as e:
            print(f"❌ 拖拽功能设置失败: {e}")
            self.drag_drop_working = False

    def _split_dnd_paths(self, raw):
        try:
            return [p.strip().strip('"') for p in self.root.tk.splitlist(raw)]
        except Exception:
            candidates = re.findall(r'\{([^}]*)\}|([^ \t\r\n]+)', raw or "")
            cleaned = []
            for a, b in candidates:
                p = (a or b).strip().strip('"')
                if p:
                    cleaned.append(p)
            return cleaned

    def on_drop(self, event):
        if not self.drag_drop_working:
            return
        try:
            raw_data = event.data or ""
            print(f"拖拽原始数据: {raw_data}")
            files = self._split_dnd_paths(raw_data)
            if not files:
                messagebox.showwarning("拖拽失败", "未解析到文件路径")
                return
            exts = ('.jpg','.jpeg','.png','.gif','.bmp','.tiff','.webp')
            image_files = [fp for fp in files if os.path.exists(fp) and fp.lower().endswith(exts)]
            if not image_files:
                messagebox.showwarning("拖拽失败", "没有找到有效的图片文件")
                return

            selected = self.item_tree.selection()
            if not selected:
                if messagebox.askyesno("添加项目", f"拖拽了 {len(image_files)} 张图片\n是否创建新项目？"):
                    self.add_item()
                    selected = self.item_tree.selection()
                else:
                    return

            if selected and len(image_files) == 1:
                item_id = selected[0]
                index = int(self.item_tree.item(item_id)['values'][0]) - 1
                img_file = image_files[0]
                if img_file not in self.items[index]['images']:
                    self.items[index]['images'].append(img_file)
                    self.refresh_item_list()
                    self.display_item_images(index)
                    self.update_stats()
                    self.set_status(f"✅ 拖拽添加图片: {os.path.basename(img_file)}")
                else:
                    self.set_status(f"⚠️ 图片已存在: {os.path.basename(img_file)}")
            else:
                if not self.items:
                    self.add_item()
                self.show_batch_assign_dialog(image_files)
        except Exception as e:
            print(f"❌ 拖拽处理错误: {e}")
            messagebox.showerror("拖拽错误", f"拖拽处理失败: {str(e)}")

    def batch_add_images(self):
        if not self.items:
            if messagebox.askyesno("提示", "当前没有项目\n是否创建新项目？"):
                self.add_item()
            else:
                return
        file_paths = filedialog.askopenfilenames(
            title="选择图片文件",
            filetypes=[
                ("图片文件","*.jpg *.jpeg *.png *.gif *.bmp *.tiff *.webp"),
                ("JPEG文件","*.jpg *.jpeg"),
                ("PNG文件","*.png"),
                ("所有文件","*.*")
            ]
        )
        if not file_paths:
            return
        valid, invalid = [], []
        for fp in file_paths:
            if self._validate_image_file(fp):
                valid.append(fp)
            else:
                invalid.append(fp)
        if invalid:
            names = [os.path.basename(f) for f in invalid[:5]]
            if len(invalid) > 5:
                names.append(f"...等{len(invalid)}个文件")
            messagebox.showwarning("文件验证","以下文件无效，将被跳过:\n"+"\n".join(names))
        if not valid:
            messagebox.showerror("错误","没有有效的图片文件")
            return
        print(f"批量添加：{len(valid)} 个有效文件")
        self.show_batch_assign_dialog(valid)

    def show_batch_assign_dialog(self, file_paths):
        if not file_paths:
            return
        if not self.items:
            messagebox.showwarning("错误","没有可分配的项目")
            return
        dialog = tk.Toplevel(self.root)
        dialog.title(f"批量分配 {len(file_paths)} 张图片")
        dialog.geometry("1000x700")
        dialog.transient(self.root)
        dialog.grab_set()
        self.current_batch_dialog = dialog

        main_frame = ttk.Frame(dialog); main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        title_frame = ttk.Frame(main_frame); title_frame.pack(fill=tk.X, pady=(0,10))
        ttk.Label(title_frame, text=f"为 {len(file_paths)} 张图片选择目标项目", font=('Arial',14,'bold')).pack(side=tk.LEFT)
        quick_frame = ttk.Frame(title_frame); quick_frame.pack(side=tk.RIGHT)

        self.assignments = {}
        def quick_assign_first():
            for var in self.assignments.values(): var.set(1)
            messagebox.showinfo("完成","已将所有图片分配到第一个项目")
        def quick_assign_even():
            total = len(self.items)
            for i, var in enumerate(self.assignments.values()):
                var.set((i % total)+1)
            messagebox.showinfo("完成", f"已将图片平均分配到 {total} 个项目")
        def quick_assign_selected():
            sel = self.item_tree.selection()
            if sel:
                idx = int(self.item_tree.item(sel[0])['values'][0])
                for var in self.assignments.values(): var.set(idx)
                messagebox.showinfo("完成", f"已将所有图片分配到项目 {idx}")
            else:
                messagebox.showwarning("提示","请先选择一个项目")

        ttk.Button(quick_frame, text="全部→项目1", command=quick_assign_first).pack(side=tk.LEFT, padx=2)
        ttk.Button(quick_frame, text="平均分配", command=quick_assign_even).pack(side=tk.LEFT, padx=2)
        ttk.Button(quick_frame, text="→选中项目", command=quick_assign_selected).pack(side=tk.LEFT, padx=2)

        list_frame = ttk.Frame(main_frame); list_frame.pack(fill=tk.BOTH, expand=True, pady=5)
        header_frame = ttk.Frame(list_frame); header_frame.pack(fill=tk.X, pady=(0,5))
        ttk.Label(header_frame, text="序号", width=6, anchor='center', font=('Arial',10,'bold')).pack(side=tk.LEFT, padx=2)
        ttk.Label(header_frame, text="文件名", width=35, anchor='w', font=('Arial',10,'bold')).pack(side=tk.LEFT, padx=2)
        ttk.Label(header_frame, text="分配到项目", width=12, anchor='center', font=('Arial',10,'bold')).pack(side=tk.LEFT, padx=2)
        ttk.Label(header_frame, text="项目描述", width=30, anchor='w', font=('Arial',10,'bold')).pack(side=tk.LEFT, padx=2)

        canvas = tk.Canvas(list_frame, height=400)
        scrollbar = ttk.Scrollbar(list_frame, orient="vertical", command=canvas.yview)
        scrollable = ttk.Frame(canvas)
        scrollable.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0,0), window=scrollable, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        for i, fp in enumerate(file_paths):
            row = ttk.Frame(scrollable); row.pack(fill=tk.X, padx=2, pady=2)
            ttk.Label(row, text=str(i+1), width=6, anchor='center').pack(side=tk.LEFT, padx=2)
            name = os.path.basename(fp); disp = name if len(name)<=30 else name[:27]+"..."
            ttk.Label(row, text=disp, width=35, anchor='w').pack(side=tk.LEFT, padx=2)

            var = tk.IntVar(value=1)
            options = []
            for j, item in enumerate(self.items):
                desc = item['description'][:20] + ("..." if len(item['description'])>20 else "")
                options.append(f"{j+1}. {desc}")
            combo = ttk.Combobox(row, width=12, state='readonly', values=options)
            combo.set(options[0]); combo.pack(side=tk.LEFT, padx=2)

            desc_var = tk.StringVar(); ttk.Label(row, textvariable=desc_var, width=30, anchor='w').pack(side=tk.LEFT, padx=2)
            def update_desc(_e=None, v=var, dv=desc_var, c=combo):
                try:
                    sel = c.get()
                    if sel:
                        pidx = int(sel.split('.')[0]) - 1
                        if 0 <= pidx < len(self.items):
                            v.set(pidx+1)
                            d = self.items[pidx]['description']
                            dv.set(d[:25] + ("..." if len(d)>25 else ""))
                except:
                    pass
            combo.bind('<<ComboboxSelected>>', update_desc)
            first = self.items[0]['description'] if self.items else ""
            desc_var.set(first[:25] + ("..." if len(first)>25 else ""))

            self.assignments[fp] = var

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        btns = ttk.Frame(main_frame); btns.pack(fill=tk.X, pady=(10,0))
        ttk.Label(btns, text=f"待分配: {len(file_paths)} 张图片 → {len(self.items)} 个项目", font=('Arial',10)).pack(side=tk.LEFT)

        def execute():
            try:
                succ=skip=err=0; stats={}
                for fp, v in self.assignments.items():
                    try:
                        pidx = v.get()-1
                        if 0 <= pidx < len(self.items):
                            if fp not in self.items[pidx]['images']:
                                self.items[pidx]['images'].append(fp); succ+=1
                                stats[f"项目{pidx+1}"] = stats.get(f"项目{pidx+1}",0)+1
                            else:
                                skip+=1
                        else:
                            err+=1
                    except:
                        err+=1
                dialog.destroy()
                self.refresh_item_list(); self.update_stats()
                sel = self.item_tree.selection()
                if sel:
                    idx = int(self.item_tree.item(sel[0])['values'][0]) - 1
                    if 0 <= idx < len(self.items): self.display_item_images(idx)
                if succ:
                    msg = "批量分配完成！\n\n"
                    msg += f"✅ 成功分配: {succ} 张\n"
                    if skip: msg += f"⚠️ 跳过重复: {skip} 张\n"
                    if err: msg += f"❌ 分配失败: {err} 张\n"
                    if stats:
                        msg += "\n分配详情:\n" + "\n".join([f"  {k}: {v} 张" for k,v in stats.items()])
                    messagebox.showinfo("批量分配完成", msg)
                    self.set_status(f"✅ 批量分配完成：{succ} 张成功")
                else:
                    messagebox.showwarning("分配结果","没有成功分配任何图片")
                    self.set_status("⚠️ 批量分配：无文件被分配")
            except Exception as e:
                messagebox.showerror("分配错误", f"批量分配失败:\n{str(e)}")
                self.set_status("❌ 批量分配失败")

        ttk.Button(btns, text="取消", command=dialog.destroy).pack(side=tk.RIGHT, padx=5)
        ttk.Button(btns, text="执行分配", command=execute).pack(side=tk.RIGHT, padx=5)

    def _validate_image_file(self, file_path):
        try:
            if not os.path.exists(file_path) or os.path.getsize(file_path)==0:
                return False
            if not file_path.lower().endswith(('.jpg','.jpeg','.png','.gif','.bmp','.tiff','.webp')):
                return False
            from PIL import Image
            with Image.open(file_path) as img:
                img.verify()
            return True
        except Exception as e:
            print(f"图片验证失败 {os.path.basename(file_path)}: {e}")
            return False

    def create_statusbar(self):
        self.status_frame = ttk.Frame(self.root); self.status_frame.pack(side=tk.BOTTOM, fill=tk.X)
        self.status_label = ttk.Label(self.status_frame, text="就绪"); self.status_label.pack(side=tk.LEFT, padx=5, pady=2)

        status_right = ttk.Frame(self.status_frame); status_right.pack(side=tk.RIGHT, padx=5, pady=2)
        info = [f"拖拽:{'✅' if self.drag_drop_working else '❌'}"]
        if EXCEL_AVAILABLE: info.append("Excel:✅")
        if PDF_AVAILABLE: info.append("PDF:✅")
        ttk.Label(status_right, text=" | ".join(info),
                  font=('Arial',8), foreground=('green' if self.drag_drop_working else 'orange')).pack(side=tk.RIGHT, padx=5)
        python_ver = f"Python {sys.version_info.major}.{sys.version_info.minor}"
        ttk.Label(status_right, text=python_ver, font=('Arial',8), foreground='gray').pack(side=tk.RIGHT, padx=10)
        self.update_stats()

    def update_stats(self):
        total_items = len(self.items)
        total_images = sum(len(it.get('images', [])) for it in self.items)
        self.stats_label.config(text=f"项目: {total_items} | 图片: {total_images}")

    def add_item(self):
        self.current_item_id += 1
        self.items.append({'id': self.current_item_id, 'description': f"维修项目 {len(self.items)+1}", 'images': []})
        self.refresh_item_list()
        if self.item_tree.get_children():
            last = self.item_tree.get_children()[-1]
            self.item_tree.selection_set(last); self.item_tree.focus(last)
        self.update_stats(); self.set_status("已添加新项目")

    def delete_selected_item(self):
        sel = self.item_tree.selection()
        if not sel:
            messagebox.showwarning("警告","请先选择要删除的项目"); return
        if messagebox.askyesno("确认","确定要删除选中的项目吗？"):
            indices = sorted([int(self.item_tree.item(i)['values'][0]) - 1 for i in sel], reverse=True)
            for idx in indices:
                if 0 <= idx < len(self.items): del self.items[idx]
            self.refresh_item_list(); self.clear_image_display(); self.update_stats()
            self.set_status("已删除选中项目")

    def move_item_up(self):
        sel = self.item_tree.selection()
        if not sel: return
        idx = int(self.item_tree.item(sel[0])['values'][0]) - 1
        if idx > 0:
            self.items[idx], self.items[idx-1] = self.items[idx-1], self.items[idx]
            self.refresh_item_list()
            new_item = self.item_tree.get_children()[idx-1]
            self.item_tree.selection_set(new_item); self.item_tree.focus(new_item)

    def move_item_down(self):
        sel = self.item_tree.selection()
        if not sel: return
        idx = int(self.item_tree.item(sel[0])['values'][0]) - 1
        if idx < len(self.items)-1:
            self.items[idx], self.items[idx+1] = self.items[idx+1], self.items[idx]
            self.refresh_item_list()
            new_item = self.item_tree.get_children()[idx+1]
            self.item_tree.selection_set(new_item); self.item_tree.focus(new_item)

    def refresh_item_list(self):
        for it in self.item_tree.get_children(): self.item_tree.delete(it)
        for i, item in enumerate(self.items):
            self.item_tree.insert('', tk.END, values=(i+1,
                                                      item['description'][:50]+('...' if len(item['description'])>50 else ''),
                                                      len(item.get('images', []))))
        self.update_max_images_per_row()

    def update_max_images_per_row(self):
        if self.items:
            self.max_images_per_row = max(len(it.get('images', [])) for it in self.items) or 1

    def on_item_select(self, _):
        sel = self.item_tree.selection()
        if sel:
            idx = int(self.item_tree.item(sel[0])['values'][0]) - 1
            if 0 <= idx < len(self.items):
                self.description_text.delete(1.0, tk.END)
                self.description_text.insert(1.0, self.items[idx]['description'])
                self.display_item_images(idx)

    def on_description_change(self, _):
        sel = self.item_tree.selection()
        if sel:
            idx = int(self.item_tree.item(sel[0])['values'][0]) - 1
            if 0 <= idx < len(self.items):
                self.items[idx]['description'] = self.description_text.get(1.0, tk.END).strip()
                self.refresh_item_list()
                new_item = self.item_tree.get_children()[idx]
                self.item_tree.selection_set(new_item); self.item_tree.focus(new_item)

    def edit_item_description(self, _): self.description_text.focus()

    def add_images(self):
        sel = self.item_tree.selection()
        if not sel:
            messagebox.showwarning("警告","请先选择一个项目"); return
        file_paths = filedialog.askopenfilenames(
            title="选择图片文件",
            filetypes=[("图片文件","*.jpg *.jpeg *.png *.gif *.bmp *.tiff *.webp"),("所有文件","*.*")]
        )
        if not file_paths: return
        idx = int(self.item_tree.item(sel[0])['values'][0]) - 1
        valid = 0
        for fp in file_paths:
            if self._validate_image_file(fp) and fp not in self.items[idx]['images']:
                self.items[idx]['images'].append(fp); valid += 1
        self.refresh_item_list(); self.display_item_images(idx); self.update_stats()
        self.set_status(f"已添加 {valid} 张有效图片")

    def display_item_images(self, item_index):
        if not (0 <= item_index < len(self.items)): return
        images = self.items[item_index].get('images', [])
        self.clear_image_display()
        if not images:
            ttk.Label(self.img_scroll_frame,
                      text="该项目暂无图片\n点击'添加图片'按钮或拖拽图片到此区域",
                      font=('Arial',12), foreground='gray', justify=tk.CENTER).pack(pady=50)
            return
        thumbnail_size = 200
        canvas_width = self.img_canvas.winfo_width() or 800
        cols = max(1, (canvas_width-20)//(thumbnail_size+15))
        for i, img_path in enumerate(images):
            row, col = divmod(i, cols)
            frame = ttk.Frame(self.img_scroll_frame, relief=tk.RIDGE, borderwidth=2)
            frame.grid(row=row, column=col, padx=8, pady=8, sticky='nw')
            try:
                if not os.path.exists(img_path):
                    ttk.Label(frame, text="文件不存在", foreground='red').pack(pady=20); continue
                with Image.open(img_path) as img:
                    img.thumbnail((thumbnail_size, thumbnail_size), Image.Resampling.LANCZOS)
                    photo = ImageTk.PhotoImage(img)
                lbl = ttk.Label(frame, image=photo); lbl.image = photo; lbl.pack(padx=3, pady=3)
                name = os.path.basename(img_path); name = name if len(name)<=25 else name[:22]+"..."
                ttk.Label(frame, text=name, font=('Arial',8)).pack()
                ttk.Button(frame, text="删除",
                           command=lambda p=img_path, idx=item_index: self.delete_image(p, idx)).pack(pady=2)
            except Exception as e:
                ttk.Label(frame, text=f"加载失败\n{str(e)[:20]}", foreground='red').pack(pady=20)
        self.img_scroll_frame.update_idletasks()
        self.img_canvas.configure(scrollregion=self.img_canvas.bbox("all"))

    def delete_image(self, img_path, item_index):
        if 0 <= item_index < len(self.items) and img_path in self.items[item_index]['images']:
            if messagebox.askyesno("确认", f"确定要删除这张图片吗？\n{os.path.basename(img_path)}"):
                self.items[item_index]['images'].remove(img_path)
                self.refresh_item_list(); self.display_item_images(item_index); self.update_stats()
                self.set_status("已删除图片")

    def clear_image_display(self):
        for w in self.img_scroll_frame.winfo_children(): w.destroy()

    def delete_selected_images(self):
        messagebox.showinfo("提示","请在图片下方点击删除按钮来删除单张图片")

    def clear_images(self):
        sel = self.item_tree.selection()
        if not sel:
            messagebox.showwarning("警告","请先选择一个项目"); return
        if messagebox.askyesno("确认","确定要清空当前项目的所有图片吗？"):
            idx = int(self.item_tree.item(sel[0])['values'][0]) - 1
            self.items[idx]['images'] = []
            self.refresh_item_list(); self.clear_image_display(); self.update_stats()
            self.set_status("已清空项目图片")

    # 统一读取标题（兼容中文输入法未提交）
    def get_project_title(self):
        try: self.root.update_idletasks()
        except: pass
        title = (self.title_entry.get() or "").strip()
        self.project_title.set(title)
        return title

    def export_excel(self):
        if not EXCEL_AVAILABLE:
            messagebox.showerror("错误","Excel导出功能需要安装openpyxl库\n请运行: pip install openpyxl"); return
        if not self.items:
            messagebox.showwarning("警告","没有数据可导出"); return
        if not self.get_project_title():
            if not messagebox.askyesno("标题提醒","您还没有设置项目标题！\n\n标题将显示在导出文档的顶部\n是否使用默认标题'维修检查报告'继续导出？"):
                return
        path = filedialog.asksaveasfilename(title="保存Excel文件", defaultextension=".xlsx", filetypes=[("Excel files","*.xlsx")])
        if path:
            try:
                self.set_status("正在导出Excel..."); self._export_excel_file(path)
                self.set_status(f"Excel文件已保存: {path}"); messagebox.showinfo("成功", f"Excel文件已保存到:\n{path}")
            except Exception as e:
                messagebox.showerror("错误", f"导出Excel失败: {str(e)}"); self.set_status("Excel导出失败")

    def export_pdf(self):
        if not PDF_AVAILABLE:
            messagebox.showerror("错误","PDF导出功能需要安装reportlab库\n请运行: pip install reportlab"); return
        if not self.items:
            messagebox.showwarning("警告","没有数据可导出"); return
        if not self.get_project_title():
            if not messagebox.askyesno("标题提醒","您还没有设置项目标题！\n\n标题将显示在导出文档的顶部\n是否使用默认标题'维修检查报告'继续导出？"):
                return
        path = filedialog.asksaveasfilename(title="保存PDF文件", defaultextension=".pdf", filetypes=[("PDF files","*.pdf")])
        if path:
            try:
                self.set_status("正在导出PDF..."); self._export_pdf_file(path)
                self.set_status(f"PDF文件已保存: {path}"); messagebox.showinfo("成功", f"PDF文件已保存到:\n{path}")
            except Exception as e:
                messagebox.showerror("错误", f"导出PDF失败: {str(e)}"); self.set_status("PDF导出失败")

    def _export_excel_file(self, file_path):
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "维修报告"
        final_title = self.get_project_title() or "维修检查报告"
        print(f"📋 Excel导出使用标题: '{final_title}'")
        title_cell = ws['A1']; title_cell.value = final_title
        title_cell.font = Font(size=20, bold=True, name='微软雅黑')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        subtitle_cell = ws['A2']; subtitle_cell.value = f"生成时间：{datetime.now().strftime('%Y年%m月%d日 %H:%M')}"
        subtitle_cell.font = Font(size=11, italic=True, name='微软雅黑')
        subtitle_cell.alignment = Alignment(horizontal='center')
        total_cols = 2 + self.max_images_per_row; end_col = chr(64 + total_cols)
        ws.merge_cells(f'A1:{end_col}1'); ws.merge_cells(f'A2:{end_col}2')
        headers = ['序号','维修内容描述'] + [f'图片{i+1}' for i in range(self.max_images_per_row)]
        for col, header in enumerate(headers, 1):
            c = ws.cell(row=4, column=col); c.value = header
            c.font = Font(bold=True, name='微软雅黑')
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
        temp_files = []
        try:
            for row_idx, item in enumerate(self.items, 5):
                ws.cell(row=row_idx, column=1).value = row_idx - 4
                ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal='center', vertical='center')
                desc_cell = ws.cell(row=row_idx, column=2)
                desc_cell.value = item['description']
                desc_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                desc_cell.font = Font(name='微软雅黑')
                images = item.get('images', [])
                row_max_height = 40
                for img_idx, img_path in enumerate(images[:self.max_images_per_row]):
                    col = 3 + img_idx
                    try:
                        if not os.path.exists(img_path):
                            ws.cell(row=row_idx, column=col).value = f"图片文件不存在:\n{os.path.basename(img_path)}"
                            continue
                        with Image.open(img_path) as img:
                            target_w, target_h = 1200, 900
                            r = img.width / img.height
                            if r > target_w/target_h:
                                new_w, new_h = target_w, int(target_w / r)
                            else:
                                new_h, new_w = target_h, int(target_h * r)
                            processed = img.resize((new_w, new_h), Image.Resampling.LANCZOS)
                            try:
                                processed = processed.filter(ImageFilter.UnsharpMask(radius=1.5, percent=150, threshold=3))
                            except: pass
                            tpath = os.path.join(tempfile.gettempdir(), f"excel_img_{uuid.uuid4().hex}.png")
                            processed.save(tpath, 'PNG', optimize=False, compress_level=1)
                        temp_files.append(tpath)
                        excel_img = xl_image.Image(tpath); scale = 0.32
                        excel_img.width = new_w * scale; excel_img.height = new_h * scale
                        ws.add_image(excel_img, f'{chr(64 + col)}{row_idx}')
                        row_max_height = max(row_max_height, new_h * scale * 0.8)
                    except Exception as e:
                        ws.cell(row=row_idx, column=col).value = f"图片处理失败:\n{os.path.basename(img_path)}"
                        print(f"图片处理错误 {img_path}: {e}")
                ws.row_dimensions[row_idx].height = row_max_height
            ws.column_dimensions['A'].width = 8; ws.column_dimensions['B'].width = 45
            for i in range(self.max_images_per_row):
                ws.column_dimensions[chr(67+i)].width = 52
            thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                          top=Side(style='thin'), bottom=Side(style='thin'))
            for row in range(4, len(self.items)+5):
                for col in range(1, total_cols+1):
                    ws.cell(row=row, column=col).border = thin
            wb.save(file_path)
            print(f"✅ Excel导出完成，标题: '{final_title}'")
        finally:
            def cleanup():
                for t in temp_files:
                    try:
                        if os.path.exists(t): os.unlink(t)
                    except: pass
            self.root.after(5000, cleanup)

    def _export_pdf_file(self, file_path):
        self._setup_chinese_fonts()
        doc = SimpleDocTemplate(file_path, pagesize=A4, topMargin=20*mm, bottomMargin=20*mm,
                                leftMargin=15*mm, rightMargin=15*mm)
        story = []; styles = getSampleStyleSheet()
        try:
            chinese = ParagraphStyle('Chinese', parent=styles['Normal'], fontName='Chinese',
                                     fontSize=10, leading=12, wordWrap='CJK')
            title_style = ParagraphStyle('ChineseTitle', parent=styles['Heading1'], fontName='Chinese',
                                         fontSize=20, spaceAfter=20, alignment=TA_CENTER, leading=24)
            subtitle_style = ParagraphStyle('Subtitle', parent=chinese, fontSize=11, alignment=TA_CENTER,
                                            textColor=colors.HexColor('#666666'))
        except:
            chinese = styles['Normal']; title_style = styles['Heading1']; subtitle_style = styles['Normal']
        final_title = self.get_project_title() or "维修检查报告"
        print(f"📋 PDF导出使用标题: '{final_title}'")
        story.append(Paragraph(final_title, title_style))
        story.append(Paragraph(f"生成时间：{datetime.now().strftime('%Y年%m月%d日 %H:%M')}", subtitle_style))
        story.append(Spacer(1, 30))
        has_images = any(it.get('images', []) for it in self.items)
        if not has_images:
            self._create_text_only_pdf_table(story, chinese)
        else:
            self._create_optimized_pdf_layout(story, chinese)
        doc.build(story)
        print(f"✅ PDF导出完成，标题: '{final_title}'")

    def _setup_chinese_fonts(self):
        try:
            system = platform.system(); font_registered = False
            if system == "Windows":
                paths = ["C:/Windows/Fonts/simsun.ttc","C:/Windows/Fonts/simhei.ttf","C:/Windows/Fonts/msyh.ttc"]
            elif system == "Darwin":
                paths = ["/Library/Fonts/Arial Unicode.ttf","/System/Library/Fonts/PingFang.ttc",
                         "/System/Library/Fonts/STHeiti Light.ttc"]
            else:
                paths = ["/usr/share/fonts/truetype/wqy/wqy-microhei.ttc","/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc"]
            for p in paths:
                try:
                    if os.path.exists(p):
                        pdfmetrics.registerFont(TTFont('Chinese', p)); font_registered = True; break
                except: continue
            if not font_registered:
                pdfmetrics.registerFont(TTFont('Chinese', 'Helvetica'))
        except: pass

    def _create_text_only_pdf_table(self, story, chinese):
        headers = ['序号','维修内容描述']
        data = [[Paragraph(h, chinese) for h in headers]]
        for i, it in enumerate(self.items):
            data.append([Paragraph(str(i+1), chinese), Paragraph(it['description'], chinese)])
        col_widths = [0.8*inch, 6.5*inch]
        table = Table(data, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND',(0,0),(-1,0),colors.lightgrey),
            ('TEXTCOLOR',(0,0),(-1,0),colors.black),
            ('ALIGN',(0,0),(0,-1),'CENTER'),
            ('ALIGN',(1,0),(1,-1),'LEFT'),
            ('FONTNAME',(0,0),(-1,0),'Chinese'),
            ('FONTSIZE',(0,0),(-1,0),12),
            ('FONTSIZE',(0,1),(-1,-1),10),
            ('BOTTOMPADDING',(0,0),(-1,0),12),
            ('TOPPADDING',(0,1),(-1,-1),6),
            ('BOTTOMPADDING',(0,1),(-1,-1),6),
            ('BACKGROUND',(0,1),(-1,-1),colors.white),
            ('GRID',(0,0),(-1,-1),0.5,colors.black),
            ('VALIGN',(0,0),(-1,-1),'TOP'),
        ]))
        story.append(table)

    def _create_optimized_pdf_layout(self, story, chinese):
        temp_files = []
        try:
            for idx, it in enumerate(self.items):
                title = Paragraph(f"{idx+1}. {it['description']}",
                                  ParagraphStyle('ItemTitle', parent=chinese, fontSize=12, fontName='Chinese',
                                                 spaceBefore=10, spaceAfter=10, leftIndent=0, leading=14))
                content = [title]
                images = it.get('images', [])
                if images:
                    elem = self._create_pdf_images(images, temp_files)
                    content.append(elem if elem else Paragraph("图片加载失败", chinese))
                else:
                    content.append(Paragraph("暂无图片", chinese))
                if idx == 0:
                    story.extend(content)
                else:
                    story.append(Spacer(1, 15))
                    if len(images) <= 3:
                        story.append(KeepTogether(content))
                    else:
                        story.extend(content)
                if idx < len(self.items)-1:
                    nxt = self.items[idx+1].get('images', [])
                    if len(nxt) > 4 or (idx+1) % 3 == 0:
                        story.append(PageBreak())
                    else:
                        story.append(Spacer(1, 20))
        finally:
            def cleanup():
                for t in temp_files:
                    try:
                        if os.path.exists(t): os.unlink(t)
                    except: pass
            atexit.register(cleanup)

    def _create_pdf_images(self, images, temp_files):
        try:
            if len(images) == 1:
                p = images[0]
                if os.path.exists(p):
                    t = self._process_pdf_image(p, temp_files, 150*mm, 100*mm)
                    if t:
                        img = RL_Image(t, width=150*mm, height=100*mm, kind='proportional')
                        table = Table([[img]], colWidths=[170*mm])
                        table.setStyle(TableStyle([('ALIGN',(0,0),(-1,-1),'CENTER')]))
                        return table
            else:
                cols = 2 if len(images) <= 4 else 3
                rows, row = [], []
                for i, p in enumerate(images):
                    if os.path.exists(p):
                        size = 70*mm if cols==2 else 50*mm
                        t = self._process_pdf_image(p, temp_files, size, size)
                        row.append(RL_Image(t, width=size, height=size, kind='proportional') if t else "")
                    else:
                        row.append("")
                    if len(row) >= cols or i == len(images)-1:
                        while len(row) < cols: row.append("")
                        rows.append(row); row=[]
                if rows:
                    col_w = 85*mm if cols==2 else 56*mm
                    table = Table(rows, colWidths=[col_w]*cols)
                    table.setStyle(TableStyle([
                        ('ALIGN',(0,0),(-1,-1),'CENTER'),
                        ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
                        ('LEFTPADDING',(0,0),(-1,-1),5),
                        ('RIGHTPADDING',(0,0),(-1,-1),5),
                        ('TOPPADDING',(0,0),(-1,-1),5),
                        ('BOTTOMPADDING',(0,0),(-1,-1),5),
                    ]))
                    return table
        except:
            pass
        return None

    def _process_pdf_image(self, img_path, temp_files, max_width, max_height):
        try:
            with Image.open(img_path) as img:
                max_w_px, max_h_px = int(max_width*10), int(max_height*10)
                r = img.width / img.height
                if img.width/max_w_px > img.height/max_h_px:
                    new_w, new_h = max_w_px, int(max_w_px / r)
                else:
                    new_h, new_w = max_h_px, int(max_h_px * r)
                resized = img.resize((new_w, new_h), Image.Resampling.LANCZOS)
                if resized.mode != 'RGB': resized = resized.convert('RGB')
                t = os.path.join(tempfile.gettempdir(), f"pdf_img_{uuid.uuid4().hex}.jpg")
                resized.save(t, 'JPEG', quality=92)
                temp_files.append(t); return t
        except:
            return None

    def new_project(self):
        if self.items and messagebox.askyesno("确认","当前项目未保存，确定要新建项目吗？"):
            self.items=[]; self.current_item_id=0; self.project_title.set("")
            self.refresh_item_list(); self.clear_image_display()
            self.description_text.delete(1.0, tk.END); self.thumbnail_cache.clear()
            self.update_stats(); self.set_status("已创建新项目")
        elif not self.items:
            self.project_title.set(""); self.set_status("新项目就绪")

    def save_project(self):
        path = filedialog.asksaveasfilename(title="保存项目文件", defaultextension=".json", filetypes=[("JSON files","*.json")])
        if path:
            try:
                data = {'title': self.get_project_title(), 'items': self.items,
                        'created_time': datetime.now().isoformat(),
                        'max_images_per_row': self.max_images_per_row, 'version':'1.7.4'}
                with open(path, 'w', encoding='utf-8') as f: json.dump(data, f, indent=2, ensure_ascii=False)
                self.set_status(f"项目已保存: {path}"); messagebox.showinfo("成功","项目保存成功！")
            except Exception as e:
                messagebox.showerror("错误", f"保存失败: {str(e)}")

    def open_project(self):
        path = filedialog.askopenfilename(title="打开项目文件", filetypes=[("JSON files","*.json")])
        if path:
            try:
                with open(path, 'r', encoding='utf-8') as f: data = json.load(f)
                self.thumbnail_cache.clear()
                self.project_title.set(data.get('title',''))
                self.items = data.get('items',[])
                self.max_images_per_row = data.get('max_images_per_row',1)
                self.current_item_id = max((it.get('id',0) for it in self.items), default=0)
                self.refresh_item_list(); self.clear_image_display()
                self.description_text.delete(1.0, tk.END); self.update_stats()
                self.set_status(f"项目已加载: {path} (v{data.get('version','1.0')})")
                messagebox.showinfo("成功","项目加载成功！")
            except Exception as e:
                messagebox.showerror("错误", f"加载失败: {str(e)}")

    def preview_report(self):
        if not self.items:
            messagebox.showwarning("警告","没有数据可预览"); return
        win = tk.Toplevel(self.root); win.title("报告预览"); win.geometry("1000x700"); win.transient(self.root)
        text = scrolledtext.ScrolledText(win, wrap=tk.WORD, font=('Courier',10)); text.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        title = self.get_project_title() or "维修检查报告"
        content = f"{'='*60}\n{title:^60}\n{'='*60}\n\n"
        content += f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
        content += f"项目总数: {len(self.items)}\n"
        content += f"图片总数: {sum(len(it.get('images',[])) for it in self.items)}\n"
        content += f"工具版本: v1.7.4\n\n"
        for i, it in enumerate(self.items):
            content += f"{'-'*60}\n项目 {i+1}: {it['description']}\n{'-'*60}\n"
            imgs = it.get('images',[])
            if imgs:
                content += f"包含图片 ({len(imgs)} 张):\n"
                for j, p in enumerate(imgs):
                    try:
                        size_k = os.path.getsize(p)/1024
                        with Image.open(p) as im: info = f"{im.width}×{im.height}"
                        content += f"  {j+1}. {os.path.basename(p)} ({size_k:.1f}KB, {info})\n"
                    except:
                        content += f"  {j+1}. {os.path.basename(p)} (无法读取信息)\n"
            else:
                content += "暂无图片\n"
            content += "\n"
        text.insert(tk.END, content); text.config(state=tk.DISABLED)

    def show_help(self):
        win = tk.Toplevel(self.root); win.title("使用说明"); win.geometry("700x620"); win.transient(self.root)
        text = f"""
维修单工具 v1.7.4 - 使用说明

• 拖拽：若状态为“❌”，请点击工具栏提示查看修复方案；安装 Homebrew 的 tkdnd 后通常自动可用。
• 导出：支持 Excel/PDF；请先设置“项目标题”，否则将提示使用默认标题。
"""
        widget = scrolledtext.ScrolledText(win, wrap=tk.WORD, font=('微软雅黑',9))
        widget.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        widget.insert(tk.END, text); widget.config(state=tk.DISABLED)

    def show_dnd_fix_guide(self):
        info_lib = ""
        try:
            tmp = tk.Tk(); info_lib = tmp.tk.eval('info library'); tmp.destroy()
        except: pass
        guide = f"""拖拽功能修复指南（macOS）

问题：Unable to load tkdnd library
原因：缺少或未找到 TkDND2.9 原生库。

一键方案（Homebrew）：
  brew install tkdnd
  export TKDND_LIBRARY="$(brew --prefix tkdnd)/lib/TkDND2.9"
  python3 运行本程序（从同一终端）

手工方案：
  1) 确认 Tcl 目录：{info_lib or '(未能检测)'}
  2) 下载/放置包含 pkgIndex.tcl 的 TkDND2.9 到上面目录，或设置 TKDND_LIBRARY 指向该目录

安装后重启本程序，状态栏应显示 拖拽:✅
"""
        messagebox.showinfo("拖拽修复指南", guide)

    def show_about(self):
        about_text = f"""维修单工具 v1.7.4

拖拽：{'✅ 可用' if self.drag_drop_working else '❌ 不可用'}
系统：{platform.system()} {platform.release()}
Python：{sys.version_info.major}.{sys.version_info.minor}.{sys.version_info.micro}
"""
        messagebox.showinfo("关于", about_text)

    def set_status(self, message):
        self.status_label.config(text=message)
        self.root.update_idletasks()

    def _on_mousewheel(self, event):
        try:
            if hasattr(event, 'delta'):
                self.img_canvas.yview_scroll(-1 if event.delta > 0 else 1, "units")
            elif event.num == 4:
                self.img_canvas.yview_scroll(-1, "units")
            elif event.num == 5:
                self.img_canvas.yview_scroll(1, "units")
        except: pass

    def _on_frame_configure(self, _):
        try:
            self.img_canvas.configure(scrollregion=self.img_canvas.bbox("all"))
        except: pass

    def refresh_display(self, *_):
        """修复：缺失该方法导致启动崩溃"""
        self.refresh_item_list()
        sel = self.item_tree.selection()
        if sel:
            idx = int(self.item_tree.item(sel[0])['values'][0]) - 1
            self.display_item_images(idx)
        self.update_stats()
        self.set_status("已刷新显示")

    def run(self):
        self.root.mainloop()

def main():
    missing = []
    if not EXCEL_AVAILABLE: missing.append("openpyxl (Excel导出)")
    if not PDF_AVAILABLE: missing.append("reportlab (PDF导出)")
    if not DRAG_DROP_AVAILABLE: missing.append("tkinterdnd2 (拖拽功能)")

    if missing:
        print("📋 可选功能状态:")
        for dep in missing: print(f"  ⚠️ {dep} - 未安装")
        print("\n💡 安装命令:\n  pip install openpyxl reportlab tkinterdnd2 Pillow")
        print("\n✅ 程序核心功能可正常使用")

    print("\n" + "="*50)
    print("🚀 启动维修单工具 v1.7.4")
    print("="*50)
    print("✅ 自动尝试定位并加载 TkDND2.9（拖拽所需）")
    print("✅ 修复：refresh_display 缺失导致崩溃")
    print("✅ 保持：高清图片质量与标题同步")
    print("📝 提醒：请设置项目标题！")
    print(f"🖥️  系统：{platform.system()} | Python: {sys.version_info.major}.{sys.version_info.minor}")
    print("="*50)

    try:
        app = RepairReportTool()
        app.run()
    except Exception as e:
        print(f"❌ 程序启动失败: {e}")
        import traceback; traceback.print_exc()

if __name__ == "__main__":
    main()
